#!/usr/bin/env python
# coding: utf-8

# In[1]:


# ============================================================
# CNC OVERTIME MINIMIZATION HEURISTIC
# Matematiksel modele uygun revize edilmiş versiyon
#
# OBJECTIVE : minimize Σm OTm  (toplam fazla mesai dakikası)
#
# HARD CONSTRAINTS (ihlal eden çözümler reddedilir):
#   C1 - Makine çakışması yok
#   C2 - Op10 → Op20 öncelik sırası korunur
#   C3 - Gecikme limiti aşılmaz (TARDINESS_LIMIT_DAYS)
#   C4 - Her iş tam adette üretilir (quantity conservation)
#   C5 - Setup süreleri çizelgeye dahil edilir
#   C6 - Makine-grup ataması eligible gruplar içinden yapılır
# ============================================================

import math
import random
import copy
import re
from pathlib import Path
from dataclasses import dataclass
from datetime import datetime, timedelta, time
from collections import defaultdict

import pandas as pd
import numpy as np

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import plotly.graph_objects as go

# ============================================================
# 0. CONFIGURATION
# ============================================================

BASE_DIR = Path("C:\\Users\\emrec\\Documents\\parallel_scheduling\\heuristic6")


def find_input_file(patterns, description):
    for pattern in patterns:
        matches = sorted(BASE_DIR.glob(pattern))
        if matches:
            return matches[0]
    available = "\n".join([p.name for p in sorted(BASE_DIR.glob("*.xlsx"))])
    raise FileNotFoundError(
        f"Could not find {description}.\n"
        f"Searched patterns: {patterns}\n"
        f"Available Excel files:\n{available}"
    )


SHIPMENT_FILE = find_input_file(
    ["492-güncel sevkiyat*.xlsx", "492*güncel*sevkiyat*.xlsx", "*sevkiyat*.xlsx"],
    "shipment file"
)
SDST_FILE = find_input_file(["SDST*.xlsx", "*SDST*.xlsx"], "SDST file")
MACHINE_GROUP_FILE = find_input_file(
    ["machine_group_data*.xlsx", "*machine*group*.xlsx"],
    "machine group file"
)

OUTPUT_FILE = BASE_DIR / "optimized_revised_model.xlsx"
OUTPUT_VERIFICATION_FILE = BASE_DIR / "heuristic_constraint_verification_revised.xlsx"
OUTPUT_GANTT_HTML_FILE = BASE_DIR / "heuristic_gantt_revised.html"

print("Using files:")
print("SHIPMENT_FILE      =", SHIPMENT_FILE)
print("SDST_FILE          =", SDST_FILE)
print("MACHINE_GROUP_FILE =", MACHINE_GROUP_FILE)

RANDOM_SEED = 42
random.seed(RANDOM_SEED)
np.random.seed(RANDOM_SEED)

PLANNING_START_HOUR = 7
DUE_TIME_HOUR = 17
OVERTIME_START = time(17, 0)
OVERTIME_END   = time(21, 0)
SHIPMENT_STYLE_OVERTIME = True

# ============================================================
# TEK AYARLANACAK PARAMETRE:
# Kaç güne kadar geç teslim kabul edilebilir? (Hard constraint)
# 0.5 = 12 saat | 1.0 = 1 gün | 1.5 = 1.5 gün | 2.0 = 2 gün
TARDINESS_LIMIT_DAYS = 1.5
# ============================================================

# Split settings
ALLOW_JOB_SPLITTING  = True
MAX_SPLITS           = 2
MIN_SPLIT_QTY        = 100

# Setup settings
INITIAL_SETUP_MIN          = 10
SAME_DIAM_SETUP_MIN        = 5
DEFAULT_DIFF_DIAM_SETUP_MIN = 20

# ============================================================
# OBJECTIVE FUNCTION
# Matematiksel model: minimize Σm OTm
#
# Amaç fonksiyonu (tek terim):
#   f(s) = Σm OTm(s)
#
# Makespan ve setup amaç fonksiyonunda YOK.
# Bunlar sadece raporlama amaçlı hesaplanır.
#
# Hard constraint ihlalleri çözümü geçersiz kılar (INFEASIBLE).
# Geçersiz çözümler INFEASIBLE_PENALTY (inf) döndürür →
# SA/LS tarafından asla kabul edilmez.
# ============================================================

# Kısıt ihlali için döndürülecek büyük değer (infeasible işaretçisi)
INFEASIBLE_PENALTY = float("inf")

# SA settings
SA_ITERATIONS      = 900
START_TEMPERATURE  = 250.0
END_TEMPERATURE    = 1.0

# Local search settings
LOCAL_SEARCH_ITERATIONS = 180


# ============================================================
# 1. HELPER FUNCTIONS
# ============================================================

def normalize_machine_name(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    s = s.replace(" ", "")
    s = s.replace(",", ".")
    s = s.replace("O", "0")
    s = s.replace("T3.", "T.3.")
    if re.match(r"^T3\d+", s):
        s = s.replace("T3", "T.3.", 1)
    return s


def excel_serial_to_datetime_date(x):
    if pd.isna(x):
        return None
    if isinstance(x, datetime):
        return x.date()
    if isinstance(x, (int, float, np.integer, np.floating)):
        return (datetime(1899, 12, 30) + timedelta(days=float(x))).date()
    parsed = pd.to_datetime(x, errors="coerce")
    if pd.isna(parsed):
        return None
    return parsed.date()


def excel_time_to_timedelta(x):
    if pd.isna(x):
        return None
    if isinstance(x, timedelta):
        return x
    if isinstance(x, datetime):
        return timedelta(hours=x.hour, minutes=x.minute, seconds=x.second)
    if isinstance(x, time):
        return timedelta(hours=x.hour, minutes=x.minute, seconds=x.second)
    if isinstance(x, (int, float, np.integer, np.floating)):
        return timedelta(days=float(x))
    parsed = pd.to_datetime(str(x), errors="coerce")
    if pd.isna(parsed):
        return None
    return timedelta(hours=parsed.hour, minutes=parsed.minute, seconds=parsed.second)


def combine_excel_date_time(date_value, time_value):
    d  = excel_serial_to_datetime_date(date_value)
    td = excel_time_to_timedelta(time_value)
    if d is None or td is None:
        return None
    return datetime.combine(d, time(0, 0)) + td


def minutes_between(start_dt, end_dt):
    if end_dt <= start_dt:
        end_dt = end_dt + timedelta(days=1)
    return (end_dt - start_dt).total_seconds() / 60.0


def overlap_minutes(a_start, a_end, b_start, b_end):
    latest_start  = max(a_start, b_start)
    earliest_end  = min(a_end,   b_end)
    return max(0.0, (earliest_end - latest_start).total_seconds() / 60.0)


def overtime_overlap_minutes(start_dt, end_dt):
    """Bir görevin fazla mesai (17:00–21:00) aralığına düşen dakikası."""
    if end_dt <= start_dt:
        return 0.0
    total = 0.0
    current_day = start_dt.date()
    last_day    = end_dt.date()
    while current_day <= last_day:
        ot_start = datetime.combine(current_day, OVERTIME_START)
        ot_end   = datetime.combine(current_day, OVERTIME_END)
        total   += overlap_minutes(start_dt, end_dt, ot_start, ot_end)
        current_day = current_day + timedelta(days=1)
    return total


def next_overtime_end(dt):
    ot_s = datetime.combine(dt.date(), OVERTIME_START)
    ot_e = datetime.combine(dt.date(), OVERTIME_END)
    if ot_s <= dt < ot_e:
        return ot_e
    return dt


def align_start_to_shipment_calendar(dt):
    if not SHIPMENT_STYLE_OVERTIME:
        return dt
    return next_overtime_end(dt)


def parse_overtime_text(value):
    if pd.isna(value):
        return 0.0
    s = str(value).strip().lower()
    if not s:
        return 0.0
    hours = 0; minutes = 0
    h = re.search(r'(\d+)\s*saat', s)
    m = re.search(r'(\d+)\s*dk',   s)
    if h: hours   = int(h.group(1))
    if m: minutes = int(m.group(1))
    if not h and not m:
        nums = re.findall(r'\d+', s)
        if nums:
            minutes = int(nums[0])
    return float(hours * 60 + minutes)


def mode_or_first(series, default=None):
    s = series.dropna()
    if len(s) == 0:
        return default
    m = s.mode()
    if len(m) > 0:
        return m.iloc[0]
    return s.iloc[0]


# ============================================================
# 2. DATA CLASSES
# ============================================================

@dataclass
class Job:
    job_id: str
    part_no: str
    due_date: datetime
    quantity: int
    diameter: float
    eligible_groups_op10: list
    eligible_groups_op20: list


@dataclass
class TaskRecord:
    job_id: str
    part_no: str
    split_id: int
    operation: int
    quantity: int
    diameter: float
    group: int
    machine: str
    start: datetime
    finish: datetime
    processing_min: float
    setup_min: float
    overtime_min: float


# ============================================================
# 3. LOAD INPUT DATA
# ============================================================

def load_shipment_operations(path):
    raw = pd.read_excel(path, sheet_name=0, header=5, usecols="D:N", engine="openpyxl")
    raw.columns = [str(c).strip().replace(":", "") for c in raw.columns]
    raw = raw.dropna(how="all")

    rename_map = {
        "Tarih": "date",
        "Başlangıç Saat": "start_time",
        "Bitiş Saat": "finish_time",
        "Parça No": "part_no",
        "Makine No": "machine",
        "CNC-1 operasyonu(piston)": "op10_flag",
        "CNC-2 operasyonu (saplama)": "op20_flag",
        "Adet": "quantity",
        "Çap": "diameter",
        "Makine Grubu": "machine_group",
        "Fazla mesai": "overtime_text",
    }
    raw = raw.rename(columns=rename_map)

    required = ["date", "start_time", "finish_time", "part_no", "machine",
                "quantity", "diameter", "machine_group"]
    for col in required:
        if col not in raw.columns:
            raise ValueError(f"Missing expected column: {col}")

    raw = raw[raw["part_no"].notna()].copy()
    raw["machine"]       = raw["machine"].apply(normalize_machine_name)
    raw["part_no"]       = raw["part_no"].astype(int).astype(str)
    raw["quantity"]      = pd.to_numeric(raw["quantity"],      errors="coerce").fillna(0).astype(int)
    raw["diameter"]      = pd.to_numeric(raw["diameter"],      errors="coerce")
    raw["machine_group"] = pd.to_numeric(raw["machine_group"], errors="coerce").astype("Int64")

    raw["operation"] = np.where(
        raw.get("op10_flag").notna(), 10,
        np.where(raw.get("op20_flag").notna(), 20, np.nan)
    )
    raw = raw[raw["operation"].notna()].copy()
    raw["operation"] = raw["operation"].astype(int)

    raw["start_dt"]    = [combine_excel_date_time(d, t) for d, t in zip(raw["date"], raw["start_time"])]
    raw["finish_dt"]   = [combine_excel_date_time(d, t) for d, t in zip(raw["date"], raw["finish_time"])]
    raw["duration_min"] = [minutes_between(s, f) for s, f in zip(raw["start_dt"], raw["finish_dt"])]
    raw["unit_min"]    = raw["duration_min"] / raw["quantity"].replace(0, np.nan)

    return raw


def load_orders_from_sayfa2(path, use_shipped_quantity=False):
    df = pd.read_excel(path, sheet_name=1, header=None, engine="openpyxl")
    jobs = []
    current_date = None
    in_block     = False

    for _, row in df.iterrows():
        first_values = row.dropna().tolist()
        if len(first_values) == 0:
            continue

        possible_date = None
        for val in row.tolist():
            if pd.isna(val): continue
            if isinstance(val, (int, float, np.integer, np.floating)) and 40000 < float(val) < 60000:
                possible_date = excel_serial_to_datetime_date(val); break
            if isinstance(val, datetime):
                possible_date = val.date(); break

        if possible_date is not None:
            current_date = possible_date
            in_block = False
            continue

        row_text = " ".join([str(v).lower() for v in row.dropna().tolist()])
        if "sipariş" in row_text and "adet" in row_text:
            in_block = True
            continue

        if in_block and current_date is not None:
            part       = row.iloc[4] if len(row) > 4 else None
            order_qty  = row.iloc[5] if len(row) > 5 else None
            shipped_qty = row.iloc[6] if len(row) > 6 else None

            if pd.isna(part) or pd.isna(order_qty):
                continue

            try:
                part_no = str(int(part))
            except Exception:
                continue

            qty_source = shipped_qty if use_shipped_quantity and not pd.isna(shipped_qty) else order_qty
            qty    = int(round(float(qty_source)))
            due_dt = datetime.combine(current_date, time(DUE_TIME_HOUR, 0))
            job_id = f"{current_date.isoformat()}__{part_no}"
            jobs.append({"job_id": job_id, "part_no": part_no, "due_date": due_dt, "quantity": qty})

    orders = pd.DataFrame(jobs)
    if orders.empty:
        raise ValueError("Could not parse any order rows from Sayfa2.")

    before_n = len(orders)
    orders = (
        orders.groupby(["job_id", "part_no"], as_index=False)
        .agg(quantity=("quantity", "sum"), due_date=("due_date", "min"))
    )
    after_n = len(orders)
    if before_n != after_n:
        print(f"Duplicate order rows aggregated: {before_n} -> {after_n} unique jobs")

    return orders


def load_machine_groups(path):
    mg = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    mg.columns = [str(c).strip() for c in mg.columns]
    mg = mg.dropna(subset=["Machine_number", "Group"]).copy()
    mg["Machine_number"] = mg["Machine_number"].apply(normalize_machine_name)
    mg["Group"]          = pd.to_numeric(mg["Group"], errors="coerce").astype(int)

    group_to_machines = defaultdict(list)
    machine_to_group  = {}
    for _, r in mg.iterrows():
        m = r["Machine_number"]
        g = int(r["Group"])
        group_to_machines[g].append(m)
        machine_to_group[m] = g

    return dict(group_to_machines), machine_to_group


def load_sdst(path):
    sdst = pd.read_excel(path, sheet_name=0, engine="openpyxl")
    sdst.columns = [str(c).strip() for c in sdst.columns]
    setup = {}
    for _, r in sdst.dropna(subset=["diam_from", "to_diam", "setup_time"]).iterrows():
        d1 = float(r["diam_from"])
        d2 = float(r["to_diam"])
        setup[(d1, d2)] = float(r["setup_time"])
    return setup


# ============================================================
# 4. BUILD MODEL DATA
# ============================================================

def build_problem_data(shipment_file, sdst_file, machine_group_file, use_shipped_quantity=False):
    ops    = load_shipment_operations(shipment_file)
    orders = load_orders_from_sayfa2(shipment_file, use_shipped_quantity=use_shipped_quantity)
    group_to_machines, machine_to_group = load_machine_groups(machine_group_file)
    setup_dict = load_sdst(sdst_file)

    part_diam = ops.groupby("part_no")["diameter"].agg(lambda x: mode_or_first(x, default=0)).to_dict()

    eligible = defaultdict(lambda: defaultdict(set))
    for _, r in ops.iterrows():
        eligible[r["part_no"]][int(r["operation"])].add(int(r["machine_group"]))

    unit_time = {}
    grouped   = ops.dropna(subset=["unit_min", "machine_group"]).groupby(
        ["part_no", "operation", "machine_group"]
    )
    for key, g in grouped:
        unit_time[(str(key[0]), int(key[1]), int(key[2]))] = float(g["unit_min"].median())

    fallback_part_op = ops.dropna(subset=["unit_min"]).groupby(
        ["part_no", "operation"])["unit_min"].median().to_dict()
    fallback_op  = ops.dropna(subset=["unit_min"]).groupby("operation")["unit_min"].median().to_dict()
    global_unit  = float(ops["unit_min"].dropna().median())

    jobs       = []
    all_groups = sorted(group_to_machines.keys())

    for _, r in orders.iterrows():
        part_no = str(r["part_no"])
        q       = int(r["quantity"])
        diam    = float(part_diam.get(part_no, ops["diameter"].dropna().median()))

        eg10 = sorted(list(eligible[part_no].get(10, set(all_groups))))
        eg20 = sorted(list(eligible[part_no].get(20, set(eg10 if eg10 else all_groups))))
        if not eg10: eg10 = all_groups
        if not eg20: eg20 = eg10

        jobs.append(Job(
            job_id=r["job_id"], part_no=part_no, due_date=r["due_date"],
            quantity=q, diameter=diam,
            eligible_groups_op10=eg10, eligible_groups_op20=eg20
        ))

    ops["shipment_file_overtime_min"] = ops.get(
        "overtime_text", pd.Series(dtype=object)
    ).apply(parse_overtime_text)
    baseline_overtime_min = float(ops["shipment_file_overtime_min"].sum())

    data = {
        "ops": ops,
        "baseline_overtime_min": baseline_overtime_min,
        "orders": orders,
        "jobs": jobs,
        "group_to_machines":  group_to_machines,
        "machine_to_group":   machine_to_group,
        "setup_dict":         setup_dict,
        "unit_time":          unit_time,
        "fallback_part_op":   fallback_part_op,
        "fallback_op":        fallback_op,
        "global_unit":        global_unit,
        "planning_start": datetime.combine(
            min([j.due_date.date() for j in jobs]),
            time(PLANNING_START_HOUR, 0)
        ),
    }
    return data


def get_unit_time(data, part_no, operation, group):
    key = (str(part_no), int(operation), int(group))
    if key in data["unit_time"]:
        return data["unit_time"][key]
    key2 = (str(part_no), int(operation))
    if key2 in data["fallback_part_op"]:
        return float(data["fallback_part_op"][key2])
    if int(operation) in data["fallback_op"]:
        return float(data["fallback_op"][int(operation)])
    return data["global_unit"]


def get_setup_time(data, from_diam, to_diam):
    if from_diam is None:
        return INITIAL_SETUP_MIN
    d1 = float(from_diam); d2 = float(to_diam)
    if abs(d1 - d2) < 1e-9:
        return SAME_DIAM_SETUP_MIN
    if (d1, d2) in data["setup_dict"]:
        return data["setup_dict"][(d1, d2)]
    return DEFAULT_DIFF_DIAM_SETUP_MIN


def common_eligible_groups(job):
    common = sorted(list(set(job.eligible_groups_op10).intersection(set(job.eligible_groups_op20))))
    if common:
        return common
    return sorted(job.eligible_groups_op10)


# ============================================================
# 5. SOLUTION REPRESENTATION
# ============================================================

def make_split_quantities(q):
    if not ALLOW_JOB_SPLITTING or q < 2 * MIN_SPLIT_QTY:
        return [int(q)]
    q1 = int(math.ceil(q / 2))
    q2 = int(q - q1)
    if q2 < MIN_SPLIT_QTY:
        return [int(q)]
    return [q1, q2]


def build_initial_solution(data):
    sol = {"splits": {}, "group": {}, "m10": {}, "m20": {}, "order": []}
    for job in data["jobs"]:
        qs     = make_split_quantities(job.quantity)
        sol["splits"][job.job_id] = qs
        groups = common_eligible_groups(job)
        for sid, _qty in enumerate(qs, start=1):
            key      = (job.job_id, sid)
            g        = random.choice(groups)
            machines = data["group_to_machines"].get(g, [])
            if not machines:
                raise ValueError(f"No machine found for group {g}.")
            sol["group"][key] = g
            sol["m10"][key]   = random.choice(machines)
            sol["m20"][key]   = random.choice(machines)
            sol["order"].append(key)

    jobs_by_id = {j.job_id: j for j in data["jobs"]}
    sol["order"].sort(key=lambda k: jobs_by_id[k[0]].due_date)
    return sol


def repair_solution(sol, data):
    jobs_by_id = {j.job_id: j for j in data["jobs"]}
    new_order  = []

    for job in data["jobs"]:
        qs = sol["splits"].get(job.job_id, [job.quantity])
        qs = [int(max(0, q)) for q in qs if q > 0]
        qs = qs[:MAX_SPLITS]

        if len(qs) > 1 and any(q < MIN_SPLIT_QTY for q in qs):
            qs = [job.quantity]

        diff = job.quantity - sum(qs)
        if len(qs) == 0:
            qs = [job.quantity]
        else:
            qs[-1] += diff

        if len(qs) > 1 and qs[-1] < MIN_SPLIT_QTY:
            qs = [job.quantity]

        sol["splits"][job.job_id] = qs
        groups = common_eligible_groups(job)

        for sid, _qty in enumerate(qs, start=1):
            key = (job.job_id, sid)
            if key not in sol["group"] or sol["group"][key] not in groups:
                sol["group"][key] = random.choice(groups)
            g        = sol["group"][key]
            machines = data["group_to_machines"].get(g, [])
            if not machines:
                g = random.choice(groups)
                sol["group"][key] = g
                machines = data["group_to_machines"][g]
            if sol["m10"].get(key) not in machines:
                sol["m10"][key] = random.choice(machines)
            if sol["m20"].get(key) not in machines:
                sol["m20"][key] = random.choice(machines)
            new_order.append(key)

    valid_keys = set(new_order)
    old_order  = []
    seen       = set()
    for k in sol.get("order", []):
        if k in valid_keys and k not in seen:
            old_order.append(k)
            seen.add(k)

    missing    = [k for k in new_order if k not in seen]
    sol["order"] = old_order + missing
    sol["order"] = [
        k for k in sol["order"]
        if k[0] in sol["splits"] and 1 <= int(k[1]) <= len(sol["splits"][k[0]])
    ]
    return sol


# ============================================================
# 6. DECODER / SCHEDULER
# ============================================================

def schedule_solution(sol, data):
    """
    Çözümü çizelgeye dönüştürür.
    Hard constraint C5 (setup) ve C2 (Op10→Op20) burada uygulanır.
    """
    sol = repair_solution(copy.deepcopy(sol), data)
    jobs_by_id = {j.job_id: j for j in data["jobs"]}

    machine_available  = {}
    machine_last_diam  = {}
    for m in data["machine_to_group"].keys():
        machine_available[m] = data["planning_start"]
        machine_last_diam[m] = None

    records    = []
    finish_op10 = {}

    for key in sol["order"]:
        job_id, sid = key
        job   = jobs_by_id[job_id]
        qty   = sol["splits"][job_id][sid - 1]
        g     = sol["group"][key]
        m10   = sol["m10"][key]
        m20   = sol["m20"][key]

        # --- Op10 ---
        unit10  = get_unit_time(data, job.part_no, 10, g)
        proc10  = qty * unit10
        setup10 = get_setup_time(data, machine_last_diam.get(m10), job.diameter)
        start10 = max(machine_available[m10], data["planning_start"]) + timedelta(minutes=setup10)
        start10 = align_start_to_shipment_calendar(start10)
        finish10 = start10 + timedelta(minutes=proc10)

        records.append(TaskRecord(
            job_id=job.job_id, part_no=job.part_no, split_id=sid,
            operation=10, quantity=qty, diameter=job.diameter,
            group=g, machine=m10, start=start10, finish=finish10,
            processing_min=proc10, setup_min=setup10,
            overtime_min=overtime_overlap_minutes(start10, finish10)
        ))
        machine_available[m10] = finish10
        machine_last_diam[m10] = job.diameter
        finish_op10[key]       = finish10

        # --- Op20 (C2: Op10 bitmeden başlayamaz) ---
        unit20  = get_unit_time(data, job.part_no, 20, g)
        proc20  = qty * unit20
        setup20 = get_setup_time(data, machine_last_diam.get(m20), job.diameter)
        start20 = max(machine_available[m20], finish_op10[key]) + timedelta(minutes=setup20)
        start20 = align_start_to_shipment_calendar(start20)
        finish20 = start20 + timedelta(minutes=proc20)

        records.append(TaskRecord(
            job_id=job.job_id, part_no=job.part_no, split_id=sid,
            operation=20, quantity=qty, diameter=job.diameter,
            group=g, machine=m20, start=start20, finish=finish20,
            processing_min=proc20, setup_min=setup20,
            overtime_min=overtime_overlap_minutes(start20, finish20)
        ))
        machine_available[m20] = finish20
        machine_last_diam[m20] = job.diameter

    schedule_df = pd.DataFrame([r.__dict__ for r in records])
    if schedule_df.empty:
        raise ValueError("Schedule is empty.")
    return schedule_df


def add_split_count(schedule_df):
    df = schedule_df.copy()
    split_counts    = df.groupby("job_id")["split_id"].max().to_dict()
    df["split_count"] = df["job_id"].map(split_counts).astype(int)
    return df


def check_machine_overlaps(schedule_df, tolerance_seconds=1):
    """C1: Makine çakışma kontrolü."""
    df = schedule_df.copy()
    df["start"]  = pd.to_datetime(df["start"])
    df["finish"] = pd.to_datetime(df["finish"])
    overlaps = []
    for machine, g in df.sort_values(["machine", "start", "finish"]).groupby("machine"):
        rows = g.to_dict("records")
        for idx in range(len(rows) - 1):
            current = rows[idx]; nxt = rows[idx + 1]
            gap_sec = (nxt["start"] - current["finish"]).total_seconds()
            if gap_sec < -tolerance_seconds:
                overlaps.append({
                    "machine": machine,
                    "task_1": f"{current['part_no']} S{current['split_id']} Op{current['operation']}",
                    "task_1_finish": current["finish"],
                    "task_2": f"{nxt['part_no']} S{nxt['split_id']} Op{nxt['operation']}",
                    "task_2_start": nxt["start"],
                    "overlap_min": round(abs(gap_sec) / 60.0, 3),
                })
    return pd.DataFrame(overlaps)


# ============================================================
# 7. FEASIBILITY CHECK  (Hard Constraints)
# ============================================================

def check_feasibility(schedule_df, sol, data):
    """
    Tüm hard kısıtları kontrol eder.

    Döndürür:
        is_feasible (bool)  — tüm kısıtlar sağlanıyor mu?
        violations  (dict)  — ihlal edilen kısıtlar ve miktarları
    """
    violations = {}

    # C1 — Makine çakışması
    overlaps = check_machine_overlaps(schedule_df)
    if not overlaps.empty:
        violations["C1_machine_overlap"] = float(overlaps["overlap_min"].sum())

    # C2 — Op10 → Op20 önceliği (schedule_solution içinde yapısal olarak
    #       garanti edilir; yine de kontrol ediyoruz)
    for (job_id, split_id), g in schedule_df.groupby(["job_id", "split_id"]):
        op10 = g[g["operation"] == 10]
        op20 = g[g["operation"] == 20]
        if op10.empty or op20.empty:
            violations.setdefault("C2_precedence", 0)
            violations["C2_precedence"] += 1
            continue
        slack = (op20["start"].iloc[0] - op10["finish"].iloc[0]).total_seconds()
        if slack < -1e-3:
            violations.setdefault("C2_precedence", 0)
            violations["C2_precedence"] += abs(slack) / 60.0

    # C3 — Gecikme limiti
    jobs_by_id          = {j.job_id: j for j in data["jobs"]}
    tardiness_limit_min = TARDINESS_LIMIT_DAYS * 24 * 60
    tard_violation      = 0.0
    schedule_df["finish"] = pd.to_datetime(schedule_df["finish"])
    for job_id, g in schedule_df[schedule_df["operation"] == 20].groupby("job_id"):
        job          = jobs_by_id[job_id]
        completion   = g["finish"].max()
        allowed      = job.due_date + timedelta(minutes=tardiness_limit_min)
        excess       = (completion - allowed).total_seconds() / 60.0
        if excess > 1e-3:
            tard_violation += excess
    if tard_violation > 0:
        violations["C3_tardiness_limit"] = tard_violation

    # C4 — Adet korunumu
    for job in data["jobs"]:
        qs = sol["splits"].get(job.job_id, [])
        if abs(sum(qs) - job.quantity) > 0:
            violations.setdefault("C4_quantity", 0)
            violations["C4_quantity"] += abs(sum(qs) - job.quantity)

    # C6 — Makine-grup ataması
    for job in data["jobs"]:
        groups = common_eligible_groups(job)
        for sid in range(1, len(sol["splits"].get(job.job_id, [])) + 1):
            key = (job.job_id, sid)
            g   = sol["group"].get(key)
            if g not in groups:
                violations.setdefault("C6_group_eligibility", 0)
                violations["C6_group_eligibility"] += 1

    is_feasible = len(violations) == 0
    return is_feasible, violations


# ============================================================
# 8. OBJECTIVE FUNCTION
#
#   minimize  f(s) = Σm OTm(s)
#
#   Eğer çözüm herhangi bir hard kısıtı ihlal ediyorsa
#   INFEASIBLE_PENALTY (inf) döndürülür → asla seçilmez.
#
#   Makespan ve setup amaç fonksiyonunda YOK.
#   Sadece raporlama için metrics dict'ine ekleniyor.
# ============================================================

def evaluate_solution(sol, data):
    """
    Döndürür:
        score   (float)  — minimize edilecek değer (inf = infeasible)
        metrics (dict)   — detaylı metrikler
    """
    try:
        sched = schedule_solution(sol, data)
        sched = add_split_count(sched)
    except Exception as e:
        return INFEASIBLE_PENALTY, {"score": INFEASIBLE_PENALTY, "infeasible_reason": str(e)}

    # Hard constraint kontrolü
    is_feasible, violations = check_feasibility(sched, sol, data)

    if not is_feasible:
        return INFEASIBLE_PENALTY, {
            "score": INFEASIBLE_PENALTY,
            "feasible": False,
            "violations": violations,
            "total_overtime": float(sched["overtime_min"].sum()),
        }

    # --------------------------------------------------------
    # Amaç fonksiyonu: f(s) = Σm OTm
    # Matematiksel modelle birebir aynı — tek terim, sadece
    # toplam fazla mesai dakikası. Başka hiçbir şey yok.
    # --------------------------------------------------------
    total_overtime = float(sched["overtime_min"].sum())   # Σm OTm

    score = total_overtime   # f(s) = Σm OTm

    # Raporlama için hesaplanan ek metrikler
    # (amaç fonksiyonunda kullanılmıyor)
    makespan    = (sched["finish"].max() - data["planning_start"]).total_seconds() / 60.0
    total_setup = float(sched["setup_min"].sum())

    # Raporlama için gecikme metrikleri (kısıt olarak kontrol edildi,
    # artık amaç fonksiyonunda yok)
    jobs_by_id = {j.job_id: j for j in data["jobs"]}
    tardiness_limit_min  = TARDINESS_LIMIT_DAYS * 24 * 60
    total_tardiness      = 0.0
    max_tardiness_days   = 0.0

    for job_id, g in sched[sched["operation"] == 20].groupby("job_id"):
        job         = jobs_by_id[job_id]
        finish_dt   = g["finish"].max()
        tard_min    = max(0.0, (finish_dt - job.due_date).total_seconds() / 60.0)
        total_tardiness   += tard_min
        max_tardiness_days = max(max_tardiness_days, tard_min / 1440.0)

    metrics = {
        "score":                    score,
        "feasible":                 True,
        "violations":               {},
        # --- Ana hedef ---
        "total_overtime":           total_overtime,
        "baseline_overtime_min":    data.get("baseline_overtime_min", 0.0),
        "overtime_improvement_min": data.get("baseline_overtime_min", 0.0) - total_overtime,
        # --- Gecikme (bilgi amaçlı, kısıt olarak kontrol edildi) ---
        "total_tardiness_from_due": total_tardiness,
        "max_tardiness_days":       max_tardiness_days,
        "tardiness_limit_days":     TARDINESS_LIMIT_DAYS,
        "total_tardiness_limit_violation": 0.0,   # Feasible ise 0
        # --- Diğer ---
        "makespan":                 makespan,
        "total_setup":              total_setup,
        "total_overlap_min":        0.0,           # Feasible ise 0
        "overlap_count":            0,
    }
    return score, metrics


# ============================================================
# 9. NEIGHBORHOOD MOVES
# ============================================================

def get_overtime_task_candidates(sol, data, top_n=8):
    try:
        sched = schedule_solution(sol, data)
    except Exception:
        return []
    if sched.empty or "overtime_min" not in sched.columns:
        return []
    ot = sched[sched["overtime_min"] > 1e-6].copy()
    if ot.empty:
        return []
    ot = ot.sort_values("overtime_min", ascending=False).head(top_n)
    return [
        {
            "key":         (r["job_id"], int(r["split_id"])),
            "operation":   int(r["operation"]),
            "machine":     r["machine"],
            "overtime_min": float(r["overtime_min"]),
            "start":       r["start"],
            "finish":      r["finish"],
            "part_no":     r["part_no"],
        }
        for _, r in ot.iterrows()
    ]


def choose_overtime_candidate(sol, data):
    candidates = get_overtime_task_candidates(sol, data, top_n=6)
    if not candidates:
        return None
    weights = [max(1.0, c["overtime_min"]) for c in candidates]
    return random.choices(candidates, weights=weights, k=1)[0]


def move_key_to_position(order, key, new_pos):
    if key not in order:
        return order
    new_order = [k for k in order if k != key]
    new_pos   = max(0, min(int(new_pos), len(new_order)))
    new_order.insert(new_pos, key)
    return new_order


def targeted_overtime_mutation(sol, data):
    """
    Fazla mesai yapan göreve odaklanan 6 hareket.
    Her hareket hard kısıtları çiğnemez; repair_solution
    geçersiz atamaları düzeltir, evaluate_solution da
    infeasible çözümleri reddeder.
    """
    new  = copy.deepcopy(sol)
    cand = choose_overtime_candidate(new, data)
    if cand is None:
        return mutate_solution(new, data)

    key    = cand["key"]
    op     = cand["operation"]
    job_id, sid = key
    jobs_by_id  = {j.job_id: j for j in data["jobs"]}
    if job_id not in jobs_by_id:
        return repair_solution(new, data)
    job = jobs_by_id[job_id]

    move = random.choice([
        "insert_earlier",           # Sıralamayi öne al → OT'den önce bitir
        "swap_with_previous",       # Bir önceki iş ile yer değiştir
        "change_machine",           # Aynı grupta farklı makine
        "change_group",             # Farklı makine grubu
        "reduce_split_qty",         # OT yapan split miktarını azalt
        "diameter_neighbor_swap",   # Yakın çaplı iş ile yer değiştir
    ])

    if move == "insert_earlier" and key in new["order"]:
        pos = new["order"].index(key)
        if pos > 0:
            step = random.randint(1, min(pos, 5))
            new["order"] = move_key_to_position(new["order"], key, pos - step)

    elif move == "swap_with_previous" and key in new["order"]:
        pos = new["order"].index(key)
        if pos > 0:
            j = max(0, pos - random.randint(1, min(pos, 2)))
            new["order"][pos], new["order"][j] = new["order"][j], new["order"][pos]

    elif move == "change_machine":
        g        = new["group"].get(key)
        machines = data["group_to_machines"].get(g, [])
        if machines:
            if op == 10:
                cur  = new["m10"].get(key)
                alts = [m for m in machines if m != cur]
                if alts: new["m10"][key] = random.choice(alts)
            else:
                cur  = new["m20"].get(key)
                alts = [m for m in machines if m != cur]
                if alts: new["m20"][key] = random.choice(alts)

    elif move == "change_group":
        groups = common_eligible_groups(job)
        cur_g  = new["group"].get(key)
        alts   = [g for g in groups if g != cur_g and data["group_to_machines"].get(g)]
        if alts:
            g        = random.choice(alts)
            machines = data["group_to_machines"][g]
            new["group"][key] = g
            new["m10"][key]   = random.choice(machines)
            new["m20"][key]   = random.choice(machines)

    elif move == "reduce_split_qty":
        qs = list(new["splits"].get(job_id, [job.quantity]))
        if len(qs) == 2 and job.quantity >= 2 * MIN_SPLIT_QTY:
            idx   = sid - 1
            other = 1 - idx
            if 0 <= idx < 2 and qs[idx] > MIN_SPLIT_QTY:
                max_delta = min(qs[idx] - MIN_SPLIT_QTY, max(10, int(job.quantity * 0.15)))
                if max_delta > 0:
                    step  = random.choice([10, 25, 50, 75, 100])
                    delta = min(max_delta, step)
                    qs[idx]   -= delta
                    qs[other] += delta
                    if qs[0] >= MIN_SPLIT_QTY and qs[1] >= MIN_SPLIT_QTY:
                        new["splits"][job_id] = [int(qs[0]), int(qs[1])]

    elif move == "diameter_neighbor_swap" and key in new["order"]:
        pos    = new["order"].index(key)
        window = list(range(max(0, pos - 4), min(len(new["order"]), pos + 5)))
        window = [w for w in window if w != pos]
        if window:
            jpos = random.choice(window)
            new["order"][pos], new["order"][jpos] = new["order"][jpos], new["order"][pos]

    return repair_solution(new, data)


def mutate_solution(sol, data):
    """
    6 genel hareket — belirli bir göreve odaklanmadan rastgele.
    """
    new        = copy.deepcopy(sol)
    jobs_by_id = {j.job_id: j for j in data["jobs"]}

    move = random.choice([
        "swap_order",         # 2 işin sırasını değiştir
        "change_machine10",   # Op10 için farklı makine
        "change_machine20",   # Op20 için farklı makine
        "change_group",       # Farklı makine grubu
        "change_split_ratio", # Split oranını değiştir
        "toggle_split",       # Böl / birleştir
    ])

    if move == "swap_order" and len(new["order"]) >= 2:
        i, j = random.sample(range(len(new["order"])), 2)
        new["order"][i], new["order"][j] = new["order"][j], new["order"][i]

    elif move == "change_machine10" and new["order"]:
        key      = random.choice(new["order"])
        g        = new["group"][key]
        machines = data["group_to_machines"].get(g, [])
        if machines: new["m10"][key] = random.choice(machines)

    elif move == "change_machine20" and new["order"]:
        key      = random.choice(new["order"])
        g        = new["group"][key]
        machines = data["group_to_machines"].get(g, [])
        if machines: new["m20"][key] = random.choice(machines)

    elif move == "change_group" and new["order"]:
        key      = random.choice(new["order"])
        job      = jobs_by_id[key[0]]
        groups   = common_eligible_groups(job)
        if groups:
            g        = random.choice(groups)
            machines = data["group_to_machines"].get(g, [])
            if machines:
                new["group"][key] = g
                new["m10"][key]   = random.choice(machines)
                new["m20"][key]   = random.choice(machines)

    elif move == "change_split_ratio":
        job = random.choice(data["jobs"])
        qs  = new["splits"].get(job.job_id, [job.quantity])
        if len(qs) == 2 and job.quantity >= 2 * MIN_SPLIT_QTY:
            delta = random.randint(-max(1, job.quantity // 10), max(1, job.quantity // 10))
            q1 = qs[0] + delta; q2 = job.quantity - q1
            if q1 >= MIN_SPLIT_QTY and q2 >= MIN_SPLIT_QTY:
                new["splits"][job.job_id] = [q1, q2]

    elif move == "toggle_split":
        job = random.choice(data["jobs"])
        qs  = new["splits"].get(job.job_id, [job.quantity])
        if len(qs) == 1 and ALLOW_JOB_SPLITTING and job.quantity >= 2 * MIN_SPLIT_QTY:
            q1 = int(math.ceil(job.quantity / 2)); q2 = int(job.quantity - q1)
            if q2 >= MIN_SPLIT_QTY:
                new["splits"][job.job_id] = [q1, q2]
                for s in [1, 2]:
                    k        = (job.job_id, s)
                    groups   = common_eligible_groups(job)
                    g        = random.choice(groups)
                    machines = data["group_to_machines"][g]
                    new["group"][k] = g
                    new["m10"][k]   = random.choice(machines)
                    new["m20"][k]   = random.choice(machines)
        elif len(qs) == 2:
            new["splits"][job.job_id] = [job.quantity]

    return repair_solution(new, data)


# ============================================================
# 10. BEST-INSERTION LOCAL SEARCH
# ============================================================

def best_insertion_for_overtime_task(sol, data, max_candidates=4, max_trials=600):
    base       = repair_solution(copy.deepcopy(sol), data)
    base_score, base_metrics = evaluate_solution(base, data)
    best       = copy.deepcopy(base)
    best_score = base_score
    best_metrics = base_metrics

    candidates = get_overtime_task_candidates(base, data, top_n=max_candidates)
    if not candidates:
        return base, base_metrics, False

    tried = 0
    for cand in candidates:
        key = cand["key"]
        op  = int(cand["operation"])
        if key not in base.get("order", []):
            continue

        g        = base["group"].get(key)
        machines = list(data["group_to_machines"].get(g, []))
        if not machines:
            continue

        current_machine   = base["m10"].get(key) if op == 10 else base["m20"].get(key)
        machine_candidates = [current_machine] + [m for m in machines if m != current_machine]
        machine_candidates = [m for m in machine_candidates if m is not None]

        old_order_without_key = [k for k in base["order"] if k != key]
        for pos in range(len(old_order_without_key) + 1):
            if tried >= max_trials: break
            for mach in machine_candidates:
                if tried >= max_trials: break
                new = copy.deepcopy(base)
                new_order = list(old_order_without_key)
                new_order.insert(pos, key)
                new["order"] = new_order
                if op == 10: new["m10"][key] = mach
                else:        new["m20"][key] = mach
                new   = repair_solution(new, data)
                score, metrics = evaluate_solution(new, data)
                tried += 1
                # Sadece feasible ve daha iyi çözümler kabul edilir
                if score < best_score and metrics.get("feasible", False):
                    best       = new
                    best_score = score
                    best_metrics = metrics
        if tried >= max_trials:
            break

    improved = best_score < base_score - 1e-6
    return best, best_metrics, improved


def best_insertion_post_search(best, data, rounds=8):
    current       = copy.deepcopy(best)
    current_score, current_metrics = evaluate_solution(current, data)
    history       = []
    for r in range(1, rounds + 1):
        candidate, metrics, improved = best_insertion_for_overtime_task(
            current, data, max_candidates=2, max_trials=180
        )
        if not improved: break
        cand_score = metrics["score"]
        if cand_score < current_score:
            current        = candidate
            current_score  = cand_score
            current_metrics = metrics
            history.append({"iteration": r, "best_score": current_score, **current_metrics})
        else:
            break
    return current, current_metrics, pd.DataFrame(history)


def split_quantity_fine_tuning_post_search(best, data, rounds=25, top_candidates=8):
    current       = repair_solution(copy.deepcopy(best), data)
    current_score, current_metrics = evaluate_solution(current, data)
    jobs_by_id    = {j.job_id: j for j in data["jobs"]}
    history       = []
    step_sizes    = [10, 25, 50, 75, 100, 150, 200]

    for r in range(1, rounds + 1):
        improved         = False
        best_round       = copy.deepcopy(current)
        best_round_score = current_score
        best_round_metrics = current_metrics

        candidates = get_overtime_task_candidates(current, data, top_n=top_candidates)
        if not candidates: break
        candidates = sorted(candidates, key=lambda c: c["overtime_min"], reverse=True)

        for cand in candidates:
            job_id, sid = cand["key"]
            job = jobs_by_id.get(job_id)
            if job is None: continue
            qs  = list(current["splits"].get(job_id, [job.quantity]))
            if len(qs) != 2 or sid not in (1, 2): continue

            idx   = sid - 1
            other = 1 - idx

            for sign_idx, sign_other in [(-1, +1), (+1, -1)]:
                for step in step_sizes:
                    new_qs = list(qs)
                    new_qs[idx]   += sign_idx * step
                    new_qs[other] += sign_other * step
                    if new_qs[idx] < MIN_SPLIT_QTY or new_qs[other] < MIN_SPLIT_QTY: continue
                    diff = job.quantity - sum(new_qs)
                    new_qs[other] += diff
                    if new_qs[idx] < MIN_SPLIT_QTY or new_qs[other] < MIN_SPLIT_QTY: continue

                    trial = copy.deepcopy(current)
                    trial["splits"][job_id] = [int(new_qs[0]), int(new_qs[1])]
                    trial = repair_solution(trial, data)
                    score, metrics = evaluate_solution(trial, data)

                    # Sadece feasible ve daha iyi kabul et
                    if score < best_round_score - 1e-6 and metrics.get("feasible", False):
                        best_round         = trial
                        best_round_score   = score
                        best_round_metrics = metrics
                        improved           = True

        if not improved: break
        current         = best_round
        current_score   = best_round_score
        current_metrics = best_round_metrics
        history.append({"iteration": r, "best_score": current_score, **current_metrics})

    return current, current_metrics, pd.DataFrame(history)


# ============================================================
# 11. SIMULATED ANNEALING + LOCAL SEARCH
# ============================================================

def simulated_annealing(data):
    """
    SA: minimize f(s) = Σm OTm
    Infeasible çözümler (hard kısıt ihlali) asla kabul edilmez.
    """
    current       = build_initial_solution(data)
    current_score, current_metrics = evaluate_solution(current, data)

    # İlk çözüm infeasible ise feasible bir tane bul
    attempts = 0
    while current_score == INFEASIBLE_PENALTY and attempts < 500:
        current = repair_solution(build_initial_solution(data), data)
        current_score, current_metrics = evaluate_solution(current, data)
        attempts += 1

    best         = copy.deepcopy(current)
    best_score   = current_score
    best_metrics = current_metrics
    history      = []

    for it in range(1, SA_ITERATIONS + 1):
        temp = START_TEMPERATURE * ((END_TEMPERATURE / START_TEMPERATURE) ** (it / SA_ITERATIONS))

        candidate = (
            targeted_overtime_mutation(current, data)
            if random.random() < 0.65
            else mutate_solution(current, data)
        )
        cand_score, cand_metrics = evaluate_solution(candidate, data)

        # Hard constraint ihlali → asla kabul etme
        if cand_score == INFEASIBLE_PENALTY:
            continue

        delta  = cand_score - current_score
        accept = (delta < 0) or (
            current_score != INFEASIBLE_PENALTY
            and random.random() < math.exp(-delta / max(temp, 1e-9))
        )

        if accept:
            current       = candidate
            current_score = cand_score
            current_metrics = cand_metrics

        if cand_score < best_score:
            best         = copy.deepcopy(candidate)
            best_score   = cand_score
            best_metrics = cand_metrics

        if it % 100 == 0 or it == 1:
            history.append({
                "iteration":   it,
                "temperature": temp,
                "best_score":  best_score,
                **best_metrics
            })

    return best, best_metrics, pd.DataFrame(history)


def local_search(best, data):
    """
    LS: Her adımda 10 komşu dene, en iyi feasible olanı seç.
    """
    current       = copy.deepcopy(best)
    current_score, current_metrics = evaluate_solution(current, data)
    improved      = True
    it            = 0
    history       = []

    while improved and it < LOCAL_SEARCH_ITERATIONS:
        improved = False
        it      += 1

        candidate_best       = current
        candidate_best_score = current_score
        candidate_best_metrics = current_metrics

        for _ in range(10):
            candidate = (
                targeted_overtime_mutation(current, data)
                if random.random() < 0.80
                else mutate_solution(current, data)
            )
            score, metrics = evaluate_solution(candidate, data)
            if score < candidate_best_score and metrics.get("feasible", False):
                candidate_best         = candidate
                candidate_best_score   = score
                candidate_best_metrics = metrics

        if candidate_best_score < current_score:
            current         = candidate_best
            current_score   = candidate_best_score
            current_metrics = candidate_best_metrics
            improved        = True

        if it % 20 == 0 or improved:
            history.append({
                "iteration": it,
                "best_score": current_score,
                **current_metrics
            })

    return current, current_metrics, pd.DataFrame(history)


# ============================================================
# 12. OUTPUT FUNCTIONS  (orijinal ile aynı, sadece metrik isimleri güncellendi)
# ============================================================

def build_split_summary(schedule_df, data):
    df = add_split_count(schedule_df)
    jobs_by_id = {j.job_id: j for j in data["jobs"]}
    rows = []
    for job_id, g in df.groupby("job_id"):
        job = jobs_by_id[job_id]
        split_quantities = (
            g[["split_id", "quantity"]].drop_duplicates().sort_values("split_id")
            .apply(lambda r: f"S{int(r['split_id'])}={int(r['quantity'])}", axis=1).tolist()
        )
        rows.append({
            "job_id":            job_id,
            "part_no":           job.part_no,
            "original_quantity": job.quantity,
            "split_count":       int(g["split_count"].max()),
            "is_split":          "YES" if int(g["split_count"].max()) > 1 else "NO",
            "split_quantities":  ", ".join(split_quantities),
            "due_date":          job.due_date,
            "final_completion":  g[g["operation"] == 20]["finish"].max(),
        })
    return pd.DataFrame(rows).sort_values(["is_split", "due_date", "part_no"], ascending=[False, True, True])


def build_schedule_output(schedule_df):
    out = add_split_count(schedule_df).copy()
    out["start"]  = pd.to_datetime(out["start"])
    out["finish"] = pd.to_datetime(out["finish"])
    out["Tarih"]                       = out["start"].dt.date
    out["Başlangıç Saat"]              = out["start"].dt.strftime("%H:%M")
    out["Bitiş Saat"]                  = out["finish"].dt.strftime("%H:%M")
    out["Parça No"]                    = out["part_no"]
    out["Makine No"]                   = out["machine"]
    out["CNC-1 operasyonu(piston)"]    = np.where(out["operation"] == 10, "X", "")
    out["CNC-2 operasyonu (saplama)"]  = np.where(out["operation"] == 20, "X", "")
    out["Adet"]                        = out["quantity"]
    out["Çap"]                         = out["diameter"]
    out["Makine Grubu"]                = out["group"]
    out["Fazla mesai (dk)"]            = out["overtime_min"].round(1)
    out["Split"]                       = out["split_id"]
    out["Split Count"]                 = out["split_count"]
    out["Split Label"]                 = "S" + out["split_id"].astype(str) + "/" + out["split_count"].astype(str)
    out["Operation Label"]             = np.where(out["operation"] == 10, "Op10", "Op20")
    out["Setup Süresi (dk)"]           = out["setup_min"].round(1)
    out["İşlem Süresi (dk)"]           = out["processing_min"].round(1)
    out["Job ID"]                      = out["job_id"]

    cols = [
        "Tarih", "Başlangıç Saat", "Bitiş Saat", "Parça No", "Makine No",
        "CNC-1 operasyonu(piston)", "CNC-2 operasyonu (saplama)",
        "Adet", "Çap", "Makine Grubu", "Fazla mesai (dk)", "Split",
        "Split Count", "Split Label", "Operation Label",
        "Setup Süresi (dk)", "İşlem Süresi (dk)", "Job ID"
    ]
    return out[cols].sort_values(["Tarih", "Makine No", "Başlangıç Saat"]).reset_index(drop=True)


def build_machine_summary(schedule_df):
    df = schedule_df.copy()
    df["start"]  = pd.to_datetime(df["start"])
    df["finish"] = pd.to_datetime(df["finish"])
    summary = df.groupby("machine").agg(
        total_processing_min=("processing_min", "sum"),
        total_setup_min=("setup_min", "sum"),
        total_overtime_min=("overtime_min", "sum"),
        first_start=("start", "min"),
        last_finish=("finish", "max"),
        task_count=("job_id", "count"),
    ).reset_index()
    summary["total_load_min"] = summary["total_processing_min"] + summary["total_setup_min"]
    return summary.sort_values("total_overtime_min", ascending=False)


def build_job_summary(schedule_df, data):
    schedule_df          = schedule_df.copy()
    schedule_df["finish"] = pd.to_datetime(schedule_df["finish"])
    jobs_by_id           = {j.job_id: j for j in data["jobs"]}
    tardiness_limit_min  = TARDINESS_LIMIT_DAYS * 24 * 60
    comp  = schedule_df[schedule_df["operation"] == 20].groupby("job_id")["finish"].max().reset_index()
    rows  = []
    for _, r in comp.iterrows():
        job       = jobs_by_id[r["job_id"]]
        finish    = r["finish"]
        tardiness = max(0.0, (finish - job.due_date).total_seconds() / 60.0)
        rows.append({
            "job_id":                   job.job_id,
            "part_no":                  job.part_no,
            "quantity":                 job.quantity,
            "due_date":                 job.due_date,
            "completion":               finish,
            "tardiness_min":            tardiness,
            "tardiness_days":           tardiness / 1440.0,
            "allowed_latest":           job.due_date + timedelta(minutes=tardiness_limit_min),
            "tardiness_limit_violated": tardiness > tardiness_limit_min,
        })
    return pd.DataFrame(rows).sort_values(["due_date", "part_no"])


def split_processing_interval_by_overtime(start_dt, finish_dt):
    segments   = []
    if finish_dt <= start_dt: return segments
    cut_points = {start_dt, finish_dt}
    d = start_dt.date()
    while d <= finish_dt.date():
        ot_s = datetime.combine(d, OVERTIME_START)
        ot_e = datetime.combine(d, OVERTIME_END)
        if start_dt < ot_s < finish_dt: cut_points.add(ot_s)
        if start_dt < ot_e < finish_dt: cut_points.add(ot_e)
        d = d + timedelta(days=1)
    pts = sorted(cut_points)
    for a, b in zip(pts[:-1], pts[1:]):
        if b <= a: continue
        mid  = a + (b - a) / 2
        ot_s = datetime.combine(mid.date(), OVERTIME_START)
        ot_e = datetime.combine(mid.date(), OVERTIME_END)
        seg_type = "OVERTIME" if ot_s <= mid < ot_e else "NORMAL"
        segments.append((a, b, seg_type))
    return segments


def create_verification_excel(schedule_df, sol, data, metrics, output_file):
    """
    Constraint Verification Excel — her hard kısıt için ayrı sayfa.

    Sayfalar:
        0. Feasibility_Summary  — tüm kısıtların genel özeti (trafik ışığı)
        1. Objective            — amaç fonksiyonu değerleri
        2. C1_MachineOverlap    — makine çakışma kontrolü
        3. C2_Precedence        — Op10 → Op20 sırası
        4. C3_Tardiness         — gecikme limiti kontrolü
        5. C4_Quantity          — adet korunumu
        6. C5_Setup             — setup süre doğrulaması
        7. C6_GroupEligibility  — makine-grup ataması
    """
    sched = add_split_count(schedule_df.copy())
    sched["start"]  = pd.to_datetime(sched["start"])
    sched["finish"] = pd.to_datetime(sched["finish"])
    jobs_by_id      = {j.job_id: j for j in data["jobs"]}
    tard_limit_min  = TARDINESS_LIMIT_DAYS * 24 * 60

    GREEN  = PatternFill("solid", fgColor="C6EFCE")
    RED    = PatternFill("solid", fgColor="FFC7CE")
    YELLOW = PatternFill("solid", fgColor="FFEB9C")
    BLUE   = PatternFill("solid", fgColor="DEEAF1")
    HEADER = PatternFill("solid", fgColor="1F4E78")
    GRAY   = PatternFill("solid", fgColor="F2F2F2")

    def hdr_font(color="FFFFFF"): return Font(color=color, bold=True)
    def thin_border():
        t = Side(style="thin", color="D9E2F3")
        return Border(left=t, right=t, top=t, bottom=t)

    def write_sheet(ws, df, pass_col="Feasible", header_color="1F4E78"):
        df = df.copy()
        for col in df.columns:
            df[col] = df[col].apply(safe_value)
        ws.append(list(df.columns))
        for cell in ws[1]:
            cell.fill      = PatternFill("solid", fgColor=header_color)
            cell.font      = hdr_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for row in df.itertuples(index=False):
            ws.append([safe_value(v) for v in row])
        if pass_col in df.columns:
            col_idx = list(df.columns).index(pass_col) + 1
            for r in range(2, ws.max_row + 1):
                val = ws.cell(r, col_idx).value
                fill = GREEN if str(val).upper() in ("TRUE","PASS","✅") else RED
                for c in range(1, ws.max_column + 1):
                    ws.cell(r, c).fill = fill
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(35, max(12, len(str(ws.cell(1, col).value)) + 4))
        ws.freeze_panes = "A2"

    def safe_value(v):
        if v is None: return ""
        if isinstance(v, float) and math.isinf(v): return "inf"
        if isinstance(v, (dict, list, set)): return str(v)
        return v

    # ── Veri hazırlığı ────────────────────────────────────────

    # C1 — Makine çakışması
    c1_df = check_machine_overlaps(sched)
    c1_feasible = c1_df.empty
    if c1_df.empty:
        c1_df = pd.DataFrame([{"machine": "—", "task_1": "—", "task_2": "—",
                                "overlap_min": 0.0, "Feasible": "✅"}])
    else:
        c1_df["Feasible"] = c1_df["overlap_min"].apply(lambda x: "✅" if x <= 0 else "❌")

    # C2 — Öncelik (Op10 → Op20)
    c2_rows = []
    for (job_id, split_id), g in sched.groupby(["job_id", "split_id"]):
        op10 = g[g["operation"] == 10]
        op20 = g[g["operation"] == 20]
        if op10.empty or op20.empty:
            c2_rows.append({
                "job_id": job_id, "split_id": int(split_id),
                "op10_finish": "—", "op20_start": "—",
                "slack_min": None, "Feasible": "❌", "Neden": "Op10 veya Op20 eksik"
            })
            continue
        f10 = op10["finish"].iloc[0]
        s20 = op20["start"].iloc[0]
        slack = (s20 - f10).total_seconds() / 60.0
        c2_rows.append({
            "job_id":     job_id,
            "split_id":   int(split_id),
            "op10_finish": str(f10)[:16],
            "op20_start":  str(s20)[:16],
            "slack_min":   round(slack, 3),
            "Feasible":    "✅" if slack >= -1e-3 else "❌",
            "Neden":       "OK" if slack >= -1e-3 else f"Op20 Op10'dan önce başlıyor ({abs(slack):.1f} dk)"
        })
    c2_df = pd.DataFrame(c2_rows)

    # C3 — Gecikme limiti
    c3_rows = []
    for job_id, g in sched[sched["operation"] == 20].groupby("job_id"):
        job        = jobs_by_id[job_id]
        completion = g["finish"].max()
        tard       = max(0.0, (completion - job.due_date).total_seconds() / 60.0)
        allowed    = job.due_date + timedelta(minutes=tard_limit_min)
        violation  = max(0.0, (completion - allowed).total_seconds() / 60.0)
        c3_rows.append({
            "job_id":           job_id,
            "part_no":          job.part_no,
            "due_date":         str(job.due_date)[:16],
            "allowed_latest":   str(allowed)[:16],
            "completion":       str(completion)[:16],
            "tardiness_min":    round(tard, 1),
            "tardiness_days":   round(tard / 1440.0, 3),
            "violation_min":    round(violation, 1),
            "Feasible":         "✅" if violation <= 1e-3 else "❌",
        })
    c3_df = pd.DataFrame(c3_rows).sort_values("tardiness_days", ascending=False)

    # C4 — Adet korunumu
    c4_rows = []
    for job in data["jobs"]:
        qs      = sol["splits"].get(job.job_id, [])
        total_q = sum(qs)
        dev     = abs(total_q - job.quantity)
        c4_rows.append({
            "job_id":          job.job_id,
            "part_no":         job.part_no,
            "siparis_adet":    job.quantity,
            "split_miktarlar": str([int(q) for q in qs]),
            "split_toplam":    total_q,
            "sapma":           dev,
            "Feasible":        "✅" if dev == 0 else "❌",
        })
    c4_df = pd.DataFrame(c4_rows)

    # C5 — Setup doğrulaması
    c5_rows = []
    for machine, g in sched.sort_values(["machine", "start"]).groupby("machine"):
        prev_diam = None
        for _, r in g.iterrows():
            expected = get_setup_time(data, prev_diam, r["diameter"])
            actual   = float(r.get("setup_min", 0) or 0)
            dev      = abs(actual - expected)
            c5_rows.append({
                "machine":        machine,
                "job_id":         r["job_id"],
                "part_no":        r["part_no"],
                "operation":      int(r["operation"]),
                "onceki_cap":     round(prev_diam, 2) if prev_diam else "İlk iş",
                "mevcut_cap":     round(float(r["diameter"]), 2),
                "beklenen_setup": round(expected, 2),
                "gercek_setup":   round(actual, 2),
                "sapma":          round(dev, 4),
                "Feasible":       "✅" if dev <= 1e-3 else "❌",
            })
            prev_diam = r["diameter"]
    c5_df = pd.DataFrame(c5_rows)

    # C6 — Grup ataması
    c6_rows = []
    for job in data["jobs"]:
        eligible = common_eligible_groups(job)
        qs = sol["splits"].get(job.job_id, [])
        for sid in range(1, len(qs) + 1):
            key  = (job.job_id, sid)
            g    = sol["group"].get(key)
            m10  = sol["m10"].get(key)
            m20  = sol["m20"].get(key)
            g_ok = g in eligible
            m10_ok = m10 in data["group_to_machines"].get(g, [])
            m20_ok = m20 in data["group_to_machines"].get(g, [])
            c6_rows.append({
                "job_id":           job.job_id,
                "part_no":          job.part_no,
                "split_id":         sid,
                "eligible_gruplar": str(eligible),
                "atanan_grup":      g,
                "m10":              m10,
                "m20":              m20,
                "grup_uygun":       "✅" if g_ok  else "❌",
                "m10_uygun":        "✅" if m10_ok else "❌",
                "m20_uygun":        "✅" if m20_ok else "❌",
                "Feasible":         "✅" if (g_ok and m10_ok and m20_ok) else "❌",
            })
    c6_df = pd.DataFrame(c6_rows)

    # OT doğrulaması (bilgi amaçlı)
    ot_rows = []
    for _, r in sched.iterrows():
        expected = overtime_overlap_minutes(r["start"], r["finish"])
        actual   = float(r.get("overtime_min", 0) or 0)
        ot_rows.append({
            "job_id":          r["job_id"],
            "part_no":         r["part_no"],
            "split_id":        int(r["split_id"]),
            "operation":       int(r["operation"]),
            "machine":         r["machine"],
            "start":           str(r["start"])[:16],
            "finish":          str(r["finish"])[:16],
            "OTm_gercek":      round(actual,   3),
            "OTm_beklenen":    round(expected, 3),
            "sapma":           round(abs(actual - expected), 4),
            "Feasible":        "✅" if abs(actual - expected) <= 1e-3 else "❌",
        })
    ot_df = pd.DataFrame(ot_rows)

    # ── Özet tablo ───────────────────────────────────────────
    def pf(df): return "✅ PASS" if (df["Feasible"] == "✅").all() else "❌ FAIL"
    def vc(df): return int((df["Feasible"] == "❌").sum())

    summary_rows = [
        {"Kısıt": "C1 — Makine çakışması",      "Kontrol Edilen": len(c1_df), "İhlal": vc(c1_df), "Durum": pf(c1_df),  "Açıklama": "Aynı makinede eş zamanlı görev olmamalı"},
        {"Kısıt": "C2 — Op10→Op20 önceliği",    "Kontrol Edilen": len(c2_df), "İhlal": vc(c2_df), "Durum": pf(c2_df),  "Açıklama": "Op20, Op10 bitmeden başlayamaz"},
        {"Kısıt": "C3 — Gecikme limiti",         "Kontrol Edilen": len(c3_df), "İhlal": vc(c3_df), "Durum": pf(c3_df),  "Açıklama": f"Teslim ≤ {TARDINESS_LIMIT_DAYS} gün geç"},
        {"Kısıt": "C4 — Adet korunumu",          "Kontrol Edilen": len(c4_df), "İhlal": vc(c4_df), "Durum": pf(c4_df),  "Açıklama": "Split toplamı sipariş adedine eşit olmalı"},
        {"Kısıt": "C5 — Setup süre doğruluğu",   "Kontrol Edilen": len(c5_df), "İhlal": vc(c5_df), "Durum": pf(c5_df),  "Açıklama": "Setup süresi SDST tablosuna uygun olmalı"},
        {"Kısıt": "C6 — Grup ataması",           "Kontrol Edilen": len(c6_df), "İhlal": vc(c6_df), "Durum": pf(c6_df),  "Açıklama": "Makine eligible gruplar içinden seçilmeli"},
        {"Kısıt": "OT hesap doğruluğu",          "Kontrol Edilen": len(ot_df), "İhlal": vc(ot_df), "Durum": pf(ot_df),  "Açıklama": "OTm = overtime_overlap_minutes() ile birebir eşleşmeli"},
    ]
    summary_df = pd.DataFrame(summary_rows)

    all_pass = all(r["Durum"] == "✅ PASS" for r in summary_rows)

    # ── Excel yaz ────────────────────────────────────────────
    wb = Workbook()

    # Sayfa 0: Feasibility Summary
    ws0       = wb.active
    ws0.title = "Feasibility_Summary"

    ws0["A1"] = "CONSTRAINT VERIFICATION RAPORU"
    ws0["A1"].font = Font(size=14, bold=True, color="FFFFFF")
    ws0["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws0.merge_cells("A1:F1")

    ws0["A2"] = f"Amaç: minimize Σm OTm  |  Gecikme limiti: {TARDINESS_LIMIT_DAYS} gün  |  Çözüm: {'✅ FEASIBLE' if all_pass else '❌ INFEASIBLE'}"
    ws0["A2"].font = Font(italic=True, size=10)
    ws0["A2"].fill = GREEN if all_pass else RED
    ws0.merge_cells("A2:F2")

    ws0["A3"] = f"Σm OTm = {metrics['total_overtime']:.1f} dk  |  Makespan = {metrics['makespan']:.1f} dk  |  Toplam setup = {metrics['total_setup']:.1f} dk"
    ws0["A3"].font = Font(size=10, color="595959")
    ws0.merge_cells("A3:F3")
    ws0.append([])

    cols = ["Kısıt", "Kontrol Edilen", "İhlal", "Durum", "Açıklama"]
    ws0.append(cols)
    for cell in ws0[5]:
        cell.fill = HEADER; cell.font = hdr_font()
        cell.alignment = Alignment(horizontal="center")

    for row in summary_rows:
        ws0.append([row[c] for c in cols])
        r = ws0.max_row
        is_pass = row["Durum"] == "✅ PASS"
        for c in range(1, 6):
            ws0.cell(r, c).fill = GREEN if is_pass else RED
            ws0.cell(r, c).font = Font(bold=(c == 1))
        ws0.cell(r, 4).font = Font(bold=True, color="006100" if is_pass else "9C0006")

    ws0.column_dimensions["A"].width = 30
    ws0.column_dimensions["B"].width = 18
    ws0.column_dimensions["C"].width = 10
    ws0.column_dimensions["D"].width = 14
    ws0.column_dimensions["E"].width = 45

    # Sayfa 1: Objective
    ws1       = wb.create_sheet("Objective")
    ws1["A1"] = "Amaç Fonksiyonu: minimize Σm OTm"
    ws1["A1"].font = Font(size=12, bold=True, color="FFFFFF")
    ws1["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws1.merge_cells("A1:C1")
    ws1.append([])

    obj_rows = [
        ("METRİK",                          "DEĞER",                             "NOT"),
        ("Σm OTm — Amaç fonksiyonu",        f"{metrics['total_overtime']:.1f} dk",  "Ana hedef — minimize edildi"),
        ("Baseline Σm OTm",                 f"{metrics['baseline_overtime_min']:.1f} dk", "Şirket çizelgesi"),
        ("İyileşme",                        f"{metrics['overtime_improvement_min']:.1f} dk", "Baseline − Optimize"),
        ("",                                "",                                    ""),
        ("Makespan",                        f"{metrics['makespan']:.1f} dk",       "Bilgi amaçlı — objective'de yok"),
        ("Toplam setup",                    f"{metrics['total_setup']:.1f} dk",    "Bilgi amaçlı — objective'de yok"),
        ("Max gecikme",                     f"{metrics['max_tardiness_days']:.3f} gün", "Bilgi amaçlı"),
        ("C3 gecikme ihlali",               f"{metrics['total_tardiness_limit_violation']:.1f} dk", "0 = feasible ✅"),
        ("C1 çakışma",                      f"{metrics['overlap_count']} adet",    "0 = feasible ✅"),
        ("Feasible",                        "✅ EVET" if metrics.get('feasible') else "❌ HAYIR", ""),
    ]
    for i, (a, b, c) in enumerate(obj_rows, start=3):
        ws1.cell(i, 1).value = a
        ws1.cell(i, 2).value = b
        ws1.cell(i, 3).value = c
        if a == "METRİK":
            for col in range(1, 4):
                ws1.cell(i, col).fill = HEADER
                ws1.cell(i, col).font = hdr_font()
        elif "OTm" in a and "Baseline" not in a and "İyileşme" not in a:
            for col in range(1, 4):
                ws1.cell(i, col).fill = PatternFill("solid", fgColor="FFE0E0")
            ws1.cell(i, 1).font = Font(bold=True)
            ws1.cell(i, 2).font = Font(bold=True, color="C00000", size=13)
        elif a == "Feasible":
            for col in range(1, 4):
                ws1.cell(i, col).fill = GREEN if metrics.get('feasible') else RED

    ws1.column_dimensions["A"].width = 32
    ws1.column_dimensions["B"].width = 22
    ws1.column_dimensions["C"].width = 38

    # Sayfalar C1–C6 + OT
    sheets_data = [
        ("C1_MachineOverlap",  c1_df,  "C1 — Makine Çakışma Kontrolü"),
        ("C2_Precedence",      c2_df,  "C2 — Op10 → Op20 Öncelik Kontrolü"),
        ("C3_Tardiness",       c3_df,  "C3 — Gecikme Limiti Kontrolü"),
        ("C4_Quantity",        c4_df,  "C4 — Adet Korunumu"),
        ("C5_Setup",           c5_df,  "C5 — Setup Süre Doğrulaması"),
        ("C6_GroupEligibility",c6_df,  "C6 — Makine-Grup Ataması"),
        ("OT_Check",           ot_df,  "Overtime Hesap Doğrulaması"),
    ]

    for sheet_name, df, title in sheets_data:
        ws = wb.create_sheet(sheet_name)
        ws["A1"] = title
        ws["A1"].font = Font(size=12, bold=True, color="FFFFFF")
        feasible_all = (df["Feasible"] == "✅").all() if "Feasible" in df.columns else True
        ws["A1"].fill = PatternFill("solid", fgColor="375623" if feasible_all else "9C0006")
        ws.merge_cells(f"A1:{get_column_letter(len(df.columns))}1")

        ihlal_sayi = int((df["Feasible"] == "❌").sum()) if "Feasible" in df.columns else 0
        ws["A2"] = f"Kontrol edilen: {len(df)} satır  |  İhlal: {ihlal_sayi}  |  {'✅ PASS' if feasible_all else '❌ FAIL'}"
        ws["A2"].font = Font(italic=True, size=9)
        ws["A2"].fill = GREEN if feasible_all else RED
        ws.merge_cells(f"A2:{get_column_letter(len(df.columns))}2")
        ws.append([])

        write_sheet(ws, df)

    wb.save(output_file)
    print(f"  Verification raporu kaydedildi: {output_file}")

    # Özet yazdır
    print(f"\n{'='*55}")
    print(f"  CONSTRAINT VERIFICATION ÖZET")
    print(f"{'='*55}")
    for r in summary_rows:
        status = "✅ PASS" if r["Durum"] == "✅ PASS" else "❌ FAIL"
        print(f"  {r['Kısıt']:35s}  {status}  ({r['İhlal']} ihlal)")
    print(f"{'='*55}")
    print(f"  GENEL SONUÇ: {'✅ FEASIBLE' if all_pass else '❌ INFEASIBLE'}")
    print(f"{'='*55}\n")


def create_html_gantt_with_shift_and_overtime(schedule_df, data, output_file):
    df = add_split_count(schedule_df).copy()
    df["start"]     = pd.to_datetime(df["start"])
    df["finish"]    = pd.to_datetime(df["finish"])
    df["setup_min"] = pd.to_numeric(df["setup_min"], errors="coerce").fillna(0.0)

    events = []
    for machine, g in df.sort_values(["machine", "start", "finish"]).groupby("machine"):
        prev_finish = data["planning_start"]
        for _, task in g.iterrows():
            setup_min    = float(task["setup_min"])
            setup_finish = task["start"]
            setup_start  = max(prev_finish, setup_finish - timedelta(minutes=setup_min))
            if setup_min > 1e-6 and setup_finish > setup_start:
                events.append({
                    "machine": machine, "Legend": "Setup", "TaskType": "SETUP",
                    "Start": setup_start, "Finish": setup_finish,
                    "Duration_min": (setup_finish - setup_start).total_seconds() / 60.0,
                    "Label": "", "job_id": task["job_id"], "part_no": task["part_no"],
                    "split_id": int(task["split_id"]), "split_count": int(task["split_count"]),
                    "operation": int(task["operation"]),
                })
            op       = int(task["operation"])
            op_label = "Op10" if op == 10 else "Op20"
            label    = f"{task['part_no']} S{int(task['split_id'])}/{int(task['split_count'])} | {op_label}"
            for seg_start, seg_finish, seg_type in split_processing_interval_by_overtime(task["start"], task["finish"]):
                legend = f"{op_label} {'Overtime' if seg_type == 'OVERTIME' else 'Normal'}"
                events.append({
                    "machine": machine, "Legend": legend, "TaskType": "PROCESS",
                    "Start": seg_start, "Finish": seg_finish,
                    "Duration_min": (seg_finish - seg_start).total_seconds() / 60.0,
                    "Label": label, "job_id": task["job_id"], "part_no": task["part_no"],
                    "split_id": int(task["split_id"]), "split_count": int(task["split_count"]),
                    "operation": op,
                })
            prev_finish = task["finish"]

    events_df = pd.DataFrame(events)
    if events_df.empty:
        with open(output_file, "w", encoding="utf-8") as f:
            f.write("<html><body><h2>No Gantt events.</h2></body></html>")
        return

    events_df["Start_min"]    = (events_df["Start"]  - data["planning_start"]).dt.total_seconds() / 60.0
    events_df["Finish_min"]   = (events_df["Finish"] - data["planning_start"]).dt.total_seconds() / 60.0
    events_df["Duration_min"] = events_df["Finish_min"] - events_df["Start_min"]
    machine_order             = sorted(events_df["machine"].unique().tolist())

    palette = {
        "Op10 Normal": "#3b82f6", "Op10 Overtime": "#ef4444",
        "Op20 Normal": "#10b981", "Op20 Overtime": "#f97316",
        "Setup": "#111111",       "Waiting": "#d1d5db",
    }
    legend_order = ["Op10 Normal", "Op10 Overtime", "Op20 Normal", "Op20 Overtime", "Setup", "Waiting"]

    fig = go.Figure()
    for legend in legend_order:
        sub = events_df[events_df["Legend"] == legend].copy()
        if sub.empty: continue
        fig.add_trace(go.Bar(
            x=sub["Duration_min"].tolist(), y=sub["machine"].tolist(),
            base=sub["Start_min"].tolist(), orientation="h", name=legend,
            marker=dict(color=palette[legend], line=dict(color="#111111", width=0.4)),
            text=sub["Label"].tolist() if legend not in {"Setup", "Waiting"} else [""] * len(sub),
            textposition="inside", insidetextanchor="middle",
            textfont=dict(color="white", size=10),
            customdata=list(zip(
                sub["job_id"], sub["part_no"], sub["split_id"], sub["split_count"],
                sub["operation"], sub["Start"].astype(str), sub["Finish"].astype(str),
                sub["Start_min"], sub["Finish_min"], sub["Duration_min"], sub["TaskType"]
            )),
            hovertemplate=(
                "Makine: %{y}<br>Job ID: %{customdata[0]}<br>Part No: %{customdata[1]}<br>"
                "Split: S%{customdata[2]}/%{customdata[3]}<br>Operation: %{customdata[4]}<br>"
                "Task type: %{customdata[10]}<br>Start: %{customdata[5]}<br>Finish: %{customdata[6]}<br>"
                "Duration dk: %{customdata[9]:.2f}<extra></extra>"
            )
        ))

    min_dt       = events_df["Start"].min()
    max_dt       = events_df["Finish"].max()
    makespan_min = float((max_dt - data["planning_start"]).total_seconds() / 60.0)

    current_day = min_dt.date()
    while current_day <= max_dt.date():
        fs_min   = (datetime.combine(current_day, time(7, 0))          - data["planning_start"]).total_seconds() / 60.0
        fe_min   = (datetime.combine(current_day, time(17, 0))         - data["planning_start"]).total_seconds() / 60.0
        ot_s_min = (datetime.combine(current_day, OVERTIME_START)      - data["planning_start"]).total_seconds() / 60.0
        ot_e_min = (datetime.combine(current_day, OVERTIME_END)        - data["planning_start"]).total_seconds() / 60.0
        fig.add_vrect(x0=fs_min,   x1=fe_min,   fillcolor="rgba(59,130,246,0.07)", line_width=0, layer="below")
        fig.add_vrect(x0=ot_s_min, x1=ot_e_min, fillcolor="rgba(239,68,68,0.08)", line_width=0, layer="below")
        for x_val, dash in [(fs_min, "dot"), (fe_min, "dash")]:
            fig.add_shape(type="line", x0=x_val, x1=x_val, y0=0, y1=1, xref="x", yref="paper",
                          line=dict(color="#2563eb", width=1, dash=dash))
        current_day = current_day + timedelta(days=1)

    fig.add_shape(type="line", x0=makespan_min, x1=makespan_min, y0=0, y1=1,
                  xref="x", yref="paper", line=dict(width=2, dash="dash", color="#dc2626"))
    fig.add_annotation(x=makespan_min, y=1.02, xref="x", yref="paper",
                       text=f"Cmax={makespan_min:.0f} dk", showarrow=False,
                       xanchor="left", font=dict(color="#dc2626", size=11))

    fig.update_layout(
        barmode="overlay",
        title=dict(text="CNC Gantt — Heuristic (minimize Σ OTm)", x=0.5),
        xaxis_title="Zaman (dk)", yaxis_title="Makine",
        yaxis=dict(categoryorder="array", categoryarray=machine_order, autorange="reversed"),
        legend_title="Görev Türü", plot_bgcolor="white", paper_bgcolor="white",
        hovermode="closest", margin=dict(l=80, r=40, t=190, b=60),
        width=2500, height=max(650, 70 * len(machine_order)),
    )
    fig.write_html(output_file, include_plotlyjs="cdn", full_html=True)


def create_excel_output(schedule_df, sa_history, ls_history, metrics, output_file, data):
    schedule_out   = build_schedule_output(schedule_df)
    machine_summary = build_machine_summary(schedule_df)
    sched_with_cnt  = add_split_count(schedule_df)
    job_summary     = build_job_summary(sched_with_cnt, data)
    split_summary   = build_split_summary(sched_with_cnt, data)
    overlap_check   = check_machine_overlaps(sched_with_cnt)
    if overlap_check.empty:
        overlap_check = pd.DataFrame([{"status": "NO OVERLAP — Feasible", "total_overlap_min": 0}])

    wb  = Workbook()
    ws1 = wb.active;                  ws1.title = "Optimized Schedule"
    ws2 = wb.create_sheet("Machine Summary")
    ws3 = wb.create_sheet("Job Summary")
    ws4 = wb.create_sheet("SA History")
    ws6 = wb.create_sheet("Split Summary")
    ws7 = wb.create_sheet("Overlap Check")

    def safe_value(v):
        """openpyxl'in yazamadığı tipleri stringe çevirir."""
        if v is None:
            return ""
        if isinstance(v, float) and math.isinf(v):
            return "inf"
        if isinstance(v, (dict, list, set)):
            return str(v)
        if isinstance(v, bool):
            return str(v)
        return v

    def write_df(ws, df):
        # Tüm değerleri güvenli hale getir
        df = df.copy()
        for col in df.columns:
            df[col] = df[col].apply(safe_value)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append([safe_value(v) for v in row])
        hf   = PatternFill("solid", fgColor="1F4E78")
        hfnt = Font(color="FFFFFF", bold=True)
        thin = Side(style="thin", color="D9E2F3")
        for cell in ws[1]:
            cell.fill      = hf
            cell.font      = hfnt
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = Border(bottom=thin)
        for col in range(1, ws.max_column + 1):
            ws.column_dimensions[get_column_letter(col)].width = min(28, max(12, len(str(ws.cell(1, col).value)) + 4))
        ws.freeze_panes = "A2"

    write_df(ws1, schedule_out)
    write_df(ws2, machine_summary)
    write_df(ws3, job_summary)
    write_df(ws6, split_summary)
    write_df(ws7, overlap_check)

    # metrics içindeki dict/inf değerlerini temizle
    metrics_clean = {
        k: (str(v) if isinstance(v, (dict, list, set)) else
            ("inf" if isinstance(v, float) and math.isinf(v) else v))
        for k, v in metrics.items()
    }
    # Boş ayırıcı satır için None doldurulmuş dict kullan
    all_cols = list(pd.DataFrame([{"phase": "FINAL", **metrics_clean}]).columns)
    empty_row = pd.DataFrame([{c: "" for c in all_cols}])

    history_df = pd.concat([
        pd.DataFrame([{"phase": "FINAL", **metrics_clean}]),
        empty_row,
        sa_history.assign(phase="SA")   if not sa_history.empty   else pd.DataFrame(),
        ls_history.assign(phase="LOCAL_SEARCH") if not ls_history.empty else pd.DataFrame(),
    ], ignore_index=True)
    write_df(ws4, history_df)

    # KPI sayfası
    ws_kpi = wb.create_sheet("KPI")
    ws_kpi["A1"] = "OBJECTIVE FUNCTION: minimize Σm OTm"
    ws_kpi["A1"].font = Font(size=13, bold=True, color="FFFFFF")
    ws_kpi["A1"].fill = PatternFill("solid", fgColor="1F4E78")
    ws_kpi.merge_cells("A1:C1")

    kpi_rows = [
        ("--- ANA HEDEF ---", "", ""),
        ("Baseline Fazla Mesai (dk)",       metrics.get("baseline_overtime_min", 0.0),  "Σm OTm — sevkiyat verisi"),
        ("Optimize Fazla Mesai (dk)",       metrics["total_overtime"],                   "Σm OTm — heuristik çözüm"),
        ("Fazla Mesai İyileşmesi (dk)",     metrics.get("overtime_improvement_min", 0.0),"Baseline − Optimize"),
        ("Objective Score  f(s) = Σm OTm", metrics["score"],                            "= Optimize Fazla Mesai (dk)"),
        ("", "", ""),
        ("--- KISITLAR (Hard Constraints) ---", "", ""),
        ("C1 Makine Çakışması",             metrics["overlap_count"],                    "0 olmalı — ihlal → infeasible"),
        ("C2 Op10→Op20 Öncelik",            "Sağlandı",                                  "Yapısal garanti"),
        ("C3 Gecikme Limiti (gün)",         TARDINESS_LIMIT_DAYS,                        "Parametre — ihlal → infeasible"),
        ("C3 Gecikme İhlali (dk)",          metrics["total_tardiness_limit_violation"],  "0 olmalı"),
        ("C4 Adet Korunumu",                "Sağlandı",                                  "repair_solution garantisi"),
        ("C6 Grup Ataması",                 "Sağlandı",                                  "Eligible gruplar içinden"),
        ("", "", ""),
        ("--- BİLGİ AMAÇLI (Objective'de YOK) ---", "", ""),
        ("Max Gecikme (gün)",               metrics["max_tardiness_days"],               "Raporlama"),
        ("Makespan (dk)",                   metrics["makespan"],                         "Raporlama — objective'de değil"),
        ("Toplam Setup (dk)",               metrics["total_setup"],                      "Raporlama — objective'de değil"),
    ]
    for i, (name, val, note) in enumerate(kpi_rows, start=2):
        ws_kpi.cell(i, 1).value = name
        ws_kpi.cell(i, 2).value = round(float(val), 3) if isinstance(val, (int, float)) else val
        ws_kpi.cell(i, 3).value = note
        ws_kpi.cell(i, 1).font  = Font(bold=("---" in str(name)))
    ws_kpi.column_dimensions["A"].width = 35
    ws_kpi.column_dimensions["B"].width = 20
    ws_kpi.column_dimensions["C"].width = 40

    wb.save(output_file)


# ============================================================
# 13. MAIN RUN
# ============================================================

if __name__ == "__main__":
    print("=" * 60)
    print("OBJECTIVE : minimize Σm OTm  (toplam fazla mesai dakikası)")
    print("HARD CONSTRAINTS:")
    print("  C1 — Makine çakışması yok")
    print("  C2 — Op10 → Op20 öncelik sırası")
    print(f"  C3 — Gecikme ≤ {TARDINESS_LIMIT_DAYS} gün ({TARDINESS_LIMIT_DAYS*24*60:.0f} dk)")
    print("  C4 — Adet korunumu")
    print("  C5 — Setup süreleri dahil")
    print("  C6 — Eligible grup ataması")
    print("=" * 60)

    print("\nLoading data...")
    GLOBAL_DATA = build_problem_data(
        shipment_file=SHIPMENT_FILE,
        sdst_file=SDST_FILE,
        machine_group_file=MACHINE_GROUP_FILE,
        use_shipped_quantity=True
    )
    print(f"Jobs: {len(GLOBAL_DATA['jobs'])} | Machines: {len(GLOBAL_DATA['machine_to_group'])}")

    print("\n--- AŞAMA 1: Simulated Annealing (minimize Σ OTm) ---")
    best_sa, metrics_sa, sa_history = simulated_annealing(GLOBAL_DATA)
    print(f"  OT: {metrics_sa['total_overtime']:.1f} dk | Feasible: {metrics_sa.get('feasible', '?')}")

    print("\n--- AŞAMA 2: Local Search ---")
    best_ls, metrics_ls, ls_history = local_search(best_sa, GLOBAL_DATA)
    print(f"  OT: {metrics_ls['total_overtime']:.1f} dk | Feasible: {metrics_ls.get('feasible', '?')}")

    print("\n--- AŞAMA 3: Best-Insertion ---")
    best_bi, metrics_bi, bi_history = best_insertion_post_search(best_ls, GLOBAL_DATA, rounds=8)
    print(f"  OT: {metrics_bi['total_overtime']:.1f} dk | Feasible: {metrics_bi.get('feasible', '?')}")

    if not bi_history.empty:
        bi_history = bi_history.copy(); bi_history["stage"] = "best_insertion"
        ls_history = pd.concat([ls_history, bi_history], ignore_index=True, sort=False)

    print("\n--- AŞAMA 4: Split Fine-Tuning ---")
    best_sq, metrics_sq, sq_history = split_quantity_fine_tuning_post_search(
        best_bi, GLOBAL_DATA, rounds=25, top_candidates=8
    )
    print(f"  OT: {metrics_sq['total_overtime']:.1f} dk | Feasible: {metrics_sq.get('feasible', '?')}")

    if not sq_history.empty:
        sq_history = sq_history.copy(); sq_history["stage"] = "split_fine_tuning"
        ls_history = pd.concat([ls_history, sq_history], ignore_index=True, sort=False)

    final_sol     = best_sq
    final_metrics = metrics_sq

    print(f"\n{'='*60}")
    print(f"FINAL SONUÇ")
    print(f"  Σm OTm (Optimize):    {final_metrics['total_overtime']:.1f} dk")
    print(f"  Σm OTm (Baseline):    {final_metrics['baseline_overtime_min']:.1f} dk")
    print(f"  İyileşme:             {final_metrics['overtime_improvement_min']:.1f} dk")
    print(f"  C1 Çakışma:           {final_metrics['overlap_count']} adet")
    print(f"  C3 Gecikme ihlali:    {final_metrics['total_tardiness_limit_violation']:.1f} dk")
    print(f"  Max gecikme:          {final_metrics['max_tardiness_days']:.2f} gün")
    print(f"  Feasible:             {final_metrics.get('feasible', '?')}")
    print(f"{'='*60}")

    final_schedule = schedule_solution(final_sol, GLOBAL_DATA)

    print("\nCreating Excel output...")
    create_excel_output(final_schedule, sa_history, ls_history, final_metrics, OUTPUT_FILE, GLOBAL_DATA)

    print("Creating Verification output...")
    create_verification_excel(final_schedule, final_sol, GLOBAL_DATA, final_metrics, OUTPUT_VERIFICATION_FILE)

    print("Creating HTML Gantt...")
    create_html_gantt_with_shift_and_overtime(final_schedule, GLOBAL_DATA, OUTPUT_GANTT_HTML_FILE)

    print(f"\nDone. Excel        : {OUTPUT_FILE}")
    print(f"Done. Verification : {OUTPUT_VERIFICATION_FILE}")
    print(f"Done. Gantt        : {OUTPUT_GANTT_HTML_FILE}")


# In[6]:


get_ipython().user_ns["check_feasibility"] = _new_check_feasibility
get_ipython().user_ns["evaluate_solution"]  = _new_evaluate_solution

import copy, math, random, time as _t
import numpy as np

_SA_ITER = 900; _LS_ITER = 180
TARDINESS_SCENARIOS = [0.5,0.75, 1.0,1.25, 1.5, 1.75]

print("TARDINESS SENSITIVITY — warm-start ile")
tard_results = []
best_sol_so_far = None
best_ot_so_far  = float("inf")

for tard in TARDINESS_SCENARIOS:
    GLOBAL_DATA["_tardiness_limit_days"] = tard
    random.seed(42); np.random.seed(42)
    t0 = _t.perf_counter()
    print(f"  ▶ {tard} gün ... ", end="", flush=True)

    # SA
    best_sa, m_sa, _ = simulated_annealing(GLOBAL_DATA)
    best_ls, m_ls, _ = local_search(best_sa, GLOBAL_DATA)
    best_bi, m_bi, _ = best_insertion_post_search(best_ls, GLOBAL_DATA, rounds=4)
    best_sq, metrics, _ = split_quantity_fine_tuning_post_search(best_bi, GLOBAL_DATA, rounds=10, top_candidates=4)

    ot   = metrics.get("total_overtime", float("inf"))
    feas = metrics.get("feasible", False)

    # Warm-start: önceki en iyi çözüm bu limitte de feasible mi?
    if best_sol_so_far is not None and feas:
        ws_sol = repair_solution(copy.deepcopy(best_sol_so_far), GLOBAL_DATA)
        ws_score, ws_metrics = evaluate_solution(ws_sol, GLOBAL_DATA)
        if ws_metrics.get("feasible") and ws_metrics.get("total_overtime", float("inf")) < ot:
            ot = ws_metrics["total_overtime"]; metrics = ws_metrics; best_sq = ws_sol

    # Monotonluk garantisi
    if feas and ot <= best_ot_so_far + 1e-6:
        best_ot_so_far  = ot
        best_sol_so_far = copy.deepcopy(best_sq)
    elif feas and best_sol_so_far is not None:
        ot = best_ot_so_far  # önceki daha iyi, onu kullan

    elapsed = _t.perf_counter() - t0
    print(f"OT={ot:.1f} dk  {'✅' if feas else '❌'}  ({elapsed:.0f}sn)")
    tard_results.append({"tardiness_gün": tard, "OT_dk": ot, "feasible": feas})

GLOBAL_DATA["_tardiness_limit_days"] = 1.5

print(f"\n{'='*45}")
ref = next((r["OT_dk"] for r in tard_results if r["tardiness_gün"]==1.5), None)
for r in tard_results:
    ot = r["OT_dk"]; feas = "✅" if r["feasible"] else "❌"
    delta = f"({ot-ref:+.1f}dk)" if ref and not math.isinf(ot) else ""
    orig = " ← orijinal" if r["tardiness_gün"]==1.5 else ""
    print(f"  {str(r['tardiness_gün'])+' gün':10s}  {ot:8.1f} dk  {feas}  {delta}{orig}")


# In[8]:


# 1.75 günün çözümünü bul ve 2.0+ için warm-start olarak kullan
TARDINESS_SCENARIOS_2 = [1.75, 2.0, 2.5, 3.0]
best_sol_so_far = None
best_ot_so_far  = float("inf")
tard_results2 = []

for tard in TARDINESS_SCENARIOS_2:
    GLOBAL_DATA["_tardiness_limit_days"] = tard
    random.seed(42); np.random.seed(42)
    t0 = _t.perf_counter()
    print(f"  ▶ {tard} gün ... ", end="", flush=True)

    best_sa, _, _ = simulated_annealing(GLOBAL_DATA)
    best_ls, _, _ = local_search(best_sa, GLOBAL_DATA)
    best_bi, _, _ = best_insertion_post_search(best_ls, GLOBAL_DATA, rounds=4)
    best_sq, metrics, _ = split_quantity_fine_tuning_post_search(
        best_bi, GLOBAL_DATA, rounds=10, top_candidates=4)

    ot   = metrics.get("total_overtime", float("inf"))
    feas = metrics.get("feasible", False)

    if best_sol_so_far is not None and feas:
        ws_sol = repair_solution(copy.deepcopy(best_sol_so_far), GLOBAL_DATA)
        ws_score, ws_metrics = evaluate_solution(ws_sol, GLOBAL_DATA)
        if ws_metrics.get("feasible") and ws_metrics.get("total_overtime", float("inf")) < ot:
            ot = ws_metrics["total_overtime"]; metrics = ws_metrics; best_sq = ws_sol

    if feas and ot <= best_ot_so_far + 1e-6:
        best_ot_so_far = ot; best_sol_so_far = copy.deepcopy(best_sq)
    elif feas and best_sol_so_far is not None:
        ot = best_ot_so_far  # monotonluk garantisi

    elapsed = _t.perf_counter() - t0
    print(f"OT={ot:.1f} dk  {'✅' if feas else '❌'}  ({elapsed:.0f}sn)")
    tard_results2.append({"tardiness_gün": tard, "OT_dk": ot, "feasible": feas})

GLOBAL_DATA["_tardiness_limit_days"] = 1.5
print(f"\n{'='*40}")
for r in tard_results2:
    ot = r["OT_dk"]; feas = "✅" if r["feasible"] else "❌"
    print(f"  {str(r['tardiness_gün'])+' gün':10s}  {ot:8.1f} dk  {feas}")


# In[9]:


# ============================================================
# TARDINESS SENSİTİVİTY — EXCEL ÇIKTISI
# Sonuçları BASE_DIR'e kaydeder
# ============================================================

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import math

# ── Veriler — buraya kendi sonuçlarını gir ───────────────────
tard_data = [
    # (tardiness_gün, OT_dk, feasible)
    (0.50,  None,    False),   # infeasible
    (0.75,  1023.3,  True),
    (1.00,   534.3,  True),
    (1.25,   534.3,  True),
    (1.50,   232.5,  True),    # ← orijinal
    (1.75,   152.6,  True),
    (2.00,   152.6,  True),
    (2.50,   152.6,  True),
    (3.00,   152.6,  True),
]

ORIG_TARD = 1.5
ORIG_OT   = 232.5

# ── Renkler ──────────────────────────────────────────────────
BLUE_H  = PatternFill("solid", fgColor="1F4E78")
BLUE_L  = PatternFill("solid", fgColor="DEEAF1")
GREEN   = PatternFill("solid", fgColor="C6EFCE")
RED     = PatternFill("solid", fgColor="FFC7CE")
ORANGE  = PatternFill("solid", fgColor="FFEB9C")
GRAY    = PatternFill("solid", fgColor="F2F2F2")

def hdr(cell):
    cell.fill = BLUE_H
    cell.font = Font(color="FFFFFF", bold=True, name="Arial")
    cell.alignment = Alignment(horizontal="center", vertical="center")

def aln(cell, h="center"):
    cell.alignment = Alignment(horizontal=h, vertical="center")

# ── Workbook ──────────────────────────────────────────────────
wb = Workbook()

# ── Sayfa 1: Özet Tablo ──────────────────────────────────────
ws = wb.active; ws.title = "Tardiness_Sensitivity"

# Başlık
ws.merge_cells("A1:H1")
ws["A1"] = "TARDINESS SENSİTİVİTY ANALİZİ — Gecikme Limiti vs Total Overtime"
ws["A1"].font = Font(size=13, bold=True, color="FFFFFF", name="Arial")
ws["A1"].fill = BLUE_H
ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
ws.row_dimensions[1].height = 24

ws.merge_cells("A2:H2")
ws["A2"] = (f"Sabit: MAX_SPLITS=2, MIN_SPLIT_QTY=100  |  "
            f"Orijinal TARDINESS_LIMIT_DAYS={ORIG_TARD} gün → OT={ORIG_OT} dk  |  "
            f"Warm-start + monotonluk garantisi")
ws["A2"].font = Font(italic=True, size=9, color="FFFFFF", name="Arial")
ws["A2"].fill = PatternFill("solid", fgColor="2E75B6")
ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
ws.row_dimensions[2].height = 16
ws.append([])

# Kolon başlıkları
headers = [
    "Tardiness\nLimiti (gün)", "Total OT\n(dk)", "OT\n(saat)",
    "Δ OT\n(dk)", "Δ OT\n(%)", "Feasible",
    "Yorum", "Not"
]
ws.append(headers)
hrow = ws.max_row
for i in range(1, 9): hdr(ws.cell(hrow, i))
ws.row_dimensions[hrow].height = 30

# Veri satırları
for tard, ot, feas in tard_data:
    is_orig = abs(tard - ORIG_TARD) < 1e-9
    is_inf  = not feas or ot is None

    delta = (ot - ORIG_OT) if (ot is not None and not is_inf) else float("nan")
    pct   = (delta / ORIG_OT * 100) if not math.isnan(delta) else float("nan")

    if is_inf:
        yorum = "Feasible çözüm bulunamadı — limit çok dar"
    elif is_orig:
        yorum = "Orijinal parametre — referans nokta"
    elif delta < -0.1:
        yorum = f"Limit genişleyince OT {abs(delta):.1f} dk azaldı"
    elif abs(delta) < 0.1:
        yorum = "Önceki en iyi çözüm bu limitte de optimal"
    else:
        yorum = f"OT {delta:.1f} dk arttı"

    not_txt = "← orijinal" if is_orig else ("infeasible" if is_inf else "")

    row = [
        tard,
        round(ot, 1) if ot is not None else "infeasible",
        round(ot/60, 2) if ot is not None else "",
        round(delta, 1) if not math.isnan(delta) else "",
        f"{pct:+.1f}%" if not math.isnan(pct) else "",
        "✅" if feas else "❌",
        yorum,
        not_txt,
    ]
    ws.append(row)
    rn = ws.max_row
    ws.row_dimensions[rn].height = 20

    # Renklendirme
    if is_orig:
        fill = PatternFill("solid", fgColor="BDD7EE")
        font = Font(bold=True, name="Arial")
    elif is_inf:
        fill = RED
        font = Font(name="Arial")
    elif delta < -0.1:
        fill = GREEN
        font = Font(name="Arial")
    else:
        fill = GRAY
        font = Font(name="Arial")

    for c in range(1, 9):
        ws.cell(rn, c).fill = fill
        ws.cell(rn, c).font = font
        ws.cell(rn, c).alignment = Alignment(
            horizontal="left" if c in [7,8] else "center",
            vertical="center"
        )

    # Feasible ikonu rengi
    f_cell = ws.cell(rn, 6)
    f_cell.font = Font(bold=True, name="Arial",
                       color="375623" if feas else "9C0006")

# Sütun genişlikleri
for col, w in [(1,18),(2,14),(3,12),(4,14),(5,12),(6,10),(7,48),(8,14)]:
    ws.column_dimensions[get_column_letter(col)].width = w
ws.freeze_panes = "A5"

# ── Sayfa 2: Grafik Verisi ────────────────────────────────────
ws2 = wb.create_sheet("Grafik_Veri")
ws2.append(["Tardiness (gün)", "OT (dk)", "Feasible"])
for c in range(1, 4): hdr(ws2.cell(1, c))

for tard, ot, feas in tard_data:
    ws2.append([tard, round(ot, 1) if ot is not None else "", "Evet" if feas else "Hayır"])
    rn = ws2.max_row
    ws2.cell(rn, 3).fill = GREEN if feas else RED

for col, w in [(1,18),(2,14),(3,12)]:
    ws2.column_dimensions[get_column_letter(col)].width = w

# Line chart ekle
chart = LineChart()
chart.title       = "Tardiness Limiti vs Total Overtime"
chart.y_axis.title = "Σm OTm (dk)"
chart.x_axis.title = "Tardiness Limiti (gün)"
chart.style       = 10
chart.width       = 20; chart.height = 12

data_ref = Reference(ws2, min_col=2, min_row=1, max_row=ws2.max_row)
cats     = Reference(ws2, min_col=1, min_row=2, max_row=ws2.max_row)
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats)
chart.series[0].graphicalProperties.line.solidFill = "185FA5"
chart.series[0].graphicalProperties.line.width     = 25000
chart.series[0].marker.symbol = "circle"
chart.series[0].marker.size   = 6

ws2.add_chart(chart, "E2")

# ── Sayfa 3: Özet KPI ────────────────────────────────────────
ws3 = wb.create_sheet("Ozet_KPI")
ws3.merge_cells("A1:C1")
ws3["A1"] = "TARDINESS SENSİTİVİTY — KPI ÖZET"
ws3["A1"].font = Font(size=12, bold=True, color="FFFFFF", name="Arial")
ws3["A1"].fill = BLUE_H
ws3["A1"].alignment = Alignment(horizontal="center")
ws3.append([])

feasible_data = [(t,o) for t,o,f in tard_data if f and o is not None]
best_tard, best_ot = min(feasible_data, key=lambda x: x[1]) if feasible_data else (None, None)

kpis = [
    ("KPI", "Değer", "Açıklama"),
    ("Orijinal Tardiness Limiti", f"{ORIG_TARD} gün", "TARDINESS_LIMIT_DAYS = 1.5"),
    ("Orijinal Total OT",         f"{ORIG_OT} dk", "Baseline — referans"),
    ("En İyi OT",                 f"{best_ot} dk" if best_ot else "—", f"Tardiness = {best_tard} gün"),
    ("OT İyileşmesi",             f"{ORIG_OT - best_ot:.1f} dk" if best_ot else "—", "Orijinal − En İyi"),
    ("İyileşme (%)",              f"{(ORIG_OT - best_ot)/ORIG_OT*100:.1f}%" if best_ot else "—", ""),
    ("",)*3,
    ("Infeasible Senaryo",        "0.5 gün", "Limit çok dar — feasible çözüm yok"),
    ("İlk Feasible",              "0.75 gün", "1023.3 dk"),
    ("Toplam Senaryo",            str(len(tard_data)), ""),
    ("Feasible Senaryo",          str(sum(1 for _,_,f in tard_data if f)), ""),
]

for i, row in enumerate(kpis, start=3):
    for j, v in enumerate(row, 1):
        c = ws3.cell(i, j, v)
        c.alignment = Alignment(horizontal="left" if j in [1,3] else "center", vertical="center")
        if row[0] == "KPI":
            hdr(c)
        elif row[0]:
            c.font = Font(bold=(j==1), name="Arial")
    if row[0] in ["En İyi OT", "OT İyileşmesi", "İyileşme (%)"]:
        for j in range(1,4): ws3.cell(i,j).fill = GREEN
    ws3.row_dimensions[i].height = 18

for col, w in [(1,28),(2,16),(3,40)]:
    ws3.column_dimensions[get_column_letter(col)].width = w

# ── Kaydet ────────────────────────────────────────────────────
output_path = BASE_DIR / "tardiness_sensitivity.xlsx"
wb.save(output_path)
print(f"✅ Kaydedildi: {output_path}")
print(f"   Sayfa 1: Özet tablo ({len(tard_data)} senaryo)")
print(f"   Sayfa 2: Grafik verisi + line chart")
print(f"   Sayfa 3: KPI özet")


# In[ ]:




